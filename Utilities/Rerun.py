import os
import openpyxl
import pandas as pd
from DataProvider import GlobalVariables
from Utilities import ConfigReader, DirectoryCreator, ExcelProcessor
from Configuration import TestSuiteSetup
from termcolor import colored
import shutil

immediateRerun = True

EXCEL_reportFilePath = DirectoryCreator.getDirectoryPath("ExcelReport") + "/Report.xlsx"

DYNAMIC_EXCEL_REPORT_PATH = os.path.join(DirectoryCreator.getDirectoryPath("ExcelReport"), 'Report.xlsx')
RESULT_COLUMNS = ['TC Execution', 'API Val', 'DB Val', 'Portal Val', 'App Val', 'UI Val', 'ChargeSlip Val', ]


def is_previous_attempt_a_failure():
    "This function is to evaluate whether the previous run was failure. if so it will return True"
    return bool((pd.read_excel(DYNAMIC_EXCEL_REPORT_PATH)[RESULT_COLUMNS] == 'Fail').sum().sum())


def is_rerun_at_the_end_enabled():
    return ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true"


def is_rerun_at_the_end_required():
    return is_previous_attempt_a_failure() and is_rerun_at_the_end_enabled()


def rerunTestAtTheEnd():
    df_rerunTestCases = pd.DataFrame(
        pd.read_excel(EXCEL_reportFilePath))

    df_rerunTestCases.set_index("Test Case ID", inplace=True)

    exeFail = []
    apiValFail = []
    dbValFail = []
    portalValFail = []
    uiValFail = []
    appValFail = []
    lst_chargeslip_failed = []
    setOfTest = set()

    # Adding Broken status, Added on Apr 11
    setOfRerunTest = set()

    for ind in df_rerunTestCases.index:
        testCaseName = str(df_rerunTestCases['File Name'][ind]) + ".py::" + str(ind)

        if str(df_rerunTestCases['TC Execution'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                      "bool_rerun_exe_val") == "True":
            exeFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['API Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                 "bool_rerun_api_val") == "True":
            apiValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['DB Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                "bool_rerun_db_val") == "True":
            dbValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['Portal Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                    "bool_rerun_portal_val") == "True":
            portalValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['App Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                 "bool_rerun_app_val") == "True":
            appValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['UI Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                "bool_rerun_ui_val") == "True":
            uiValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))

        if str(df_rerunTestCases['ChargeSlip Val'][ind]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                "bool_rerun_chargeslip_val") == "True":
            lst_chargeslip_failed.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))

    ls_rerunTestCases = list(setOfTest)
    listToStr = ' '.join([str(elem) for elem in ls_rerunTestCases])

    if str(ConfigReader.read_config("Validations", "bool_rerun_at_the_end_parallel")).lower() == 'true':
        number_of_threads = TestSuiteSetup.calculateTestCasesCountForParallelExecution()
    else:
        number_of_threads = ""

    print("python3.8 -m pytest -v " + listToStr + " " + number_of_threads + ' --alluredir=' + DirectoryCreator.getDirectoryPath("AllureReport"))
    print(list(setOfTest))
    print("List of exe tcs failed", exeFail)
    print("List of apiVal tcs failed", apiValFail)
    print("List of dbVal tcs failed", dbValFail)
    print("List of portalVal tcs failed", portalValFail)
    print("List of appVal tcs failed", appValFail)
    print("List of uiVal tcs failed", uiValFail)
    print("List of chargeslip tcs failed", lst_chargeslip_failed)
    print(GlobalVariables.df_testCasesDetail)

    if len(listToStr) > 0:
        # To send the testcaseID as a list to change the overall_Status as empty
        ls_TestCasesForRerun = list(setOfRerunTest)
        changeOverallStatusToEmpty(ls_TestCasesForRerun)
        os.system(
            "python3.8 -m pytest -v " + listToStr + " " + number_of_threads +' --alluredir=' + DirectoryCreator.getDirectoryPath("AllureReport"))

    return len(listToStr.strip())


def isRerunRequiredImmediately(testCaseID):
    isRerunRequired = False
    df_rerunTestCases = pd.DataFrame(
        pd.read_excel(EXCEL_reportFilePath))

    df_rerunTestCases.set_index("Test Case ID", inplace=True)

    testCaseName = str(df_rerunTestCases['File Name'][testCaseID]) + ".py::" + str(testCaseID)

    if str(df_rerunTestCases['TC Execution'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                         "bool_rerun_exe_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['API Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                    "bool_rerun_api_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['DB Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                   "bool_rerun_db_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['Portal Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                       "bool_rerun_portal_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['App Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                    "bool_rerun_app_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['UI Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                   "bool_rerun_ui_val") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['ChargeSlip Val'][testCaseID]).lower() == 'fail' and ConfigReader.read_config("Validations",
                                                                                                   "bool_rerun_chargeslip_val") == "True":
        isRerunRequired = True
    print("Rerun required = " + str(isRerunRequired))
    return isRerunRequired


# To change the value of rerun testcases overall_status to empty in Report excel, so that it will set as Broken in
# case of any connectivity issues
def changeOverallStatusToEmpty(ls_TestCasesForRerun):
    wb = openpyxl.load_workbook(EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for rerun_tesecase in ls_TestCasesForRerun:
        for i in range(2, sheet.max_row + 1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            if testcase == rerun_tesecase:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = ""
    wb.save(EXCEL_reportFilePath)


def rerunTestImmediately(testCaseID, testCaseFileName, rerunCount, request):
    print("Starting the immediate rerun")
    if setRerunCount(testCaseID, rerunCount):
        # make status empty
        rerunCommand = "python3.8 -m pytest -v " + testCaseFileName + ".py::" + testCaseID + ' --alluredir=' + DirectoryCreator.getDirectoryPath(
            "AllureReport")
        print(rerunCommand)

        if rerunCount >= 0:
            GlobalVariables.time_calc.teardown.pause()
            print(colored("Teardown Timer paused (since rerun) inside rerunTestImmediately method".center(
                shutil.get_terminal_size().columns, "="), 'cyan'))

            # To send the testcaseID as a list to change the overall_Status as empty
            setOfRerunTest = set()
            setOfRerunTest.add(testCaseID)
            ls_TestCasesForRerun = list(setOfRerunTest)
            changeOverallStatusToEmpty(ls_TestCasesForRerun)

            print("$$$$$$$$$$$$$$$$$$$$ Rerun Immediately #################")
            os.system(rerunCommand)
            GlobalVariables.time_calc.teardown.resume()
            print(colored("Teardown Timer resumed (after rerun) inside rerunTestImmediately method".center(
                shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        print("Cannot perform rerun since the rerun count is 0 or the rerun sheet is not accessible.")


from DataProvider.GlobalConstants import RUNTIME_DIR

xl_RerunCountPath = os.path.join(RUNTIME_DIR, 'RerunCount.xlsx')

xl_Timestamp = str(ConfigReader.read_config_paths("System", "automation_suite_path")) + "/TestCases/Timestamp.xlsx"


def prepareImmediateRerunExcel():
    # Added on Apr 11
    df_overallTClist = pd.read_excel(
        str(ConfigReader.read_config_paths("System", "automation_suite_path")) + "/Runtime/AllTestcaseSuite.xlsx")

    df_overallTClist.set_index('Test Case ID', inplace=True)

    df_overallTClist.drop(columns=['File Name', 'Execute', 'Unnamed: 0'], inplace=True)

    # Added for adding rerun attempts in Report excel
    df_overallTClist.drop(columns=['Directory Name'], inplace=True)

    reRunCount = []
    for i in range(0, len(df_overallTClist.index)):
        try:
            configuredImmediateRerunCount = int(ConfigReader.read_config("Validations", "int_rerun_count"))
            reRunCount.append(configuredImmediateRerunCount)
        except ValueError:
            print("Configured Immediate Rerun Count is Invalid")
            return None

    df_overallTClist['Rerun Count'] = reRunCount
    df_overallTClist.to_excel(xl_RerunCountPath, sheet_name="Rerun Count")


def prepareAtTheEndRerunExcel():
    df = pd.DataFrame({'Rerun Count': [int(ConfigReader.read_config("Validations", "int_rerun_count"))]})
    print(df)
    writer = pd.ExcelWriter(xl_RerunCountPath, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Rerun At The End', index=False)
    writer.save()


def getRerunCount(testCaseID):
    if ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true":
        try:
            df_TClist = pd.read_excel(xl_RerunCountPath, sheet_name="Rerun At The End")
            try:
                return df_TClist['Rerun Count'][0]
            except:
                pass
        except:
            return -2

    if (ConfigReader.read_config("Validations", "bool_rerun_immediately").lower() == "true" and
            ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "false"):
        try:
            df_TClist = pd.read_excel(xl_RerunCountPath, sheet_name="Rerun Count")
            df_TClist.set_index('Test Case ID', inplace=True)
            try:
                return df_TClist['Rerun Count'][testCaseID]
            except:
                pass
        except:
            return -2


def set_rerun_at_the_end_count_up_to_report_excel_file(count_up_rerun):
    if count_up_rerun:
        DYNAMIC_EXCEL_REPORT_PATH = DirectoryCreator.getDirectoryPath("ExcelReport") + "/Report.xlsx"
        df = pd.read_excel(DYNAMIC_EXCEL_REPORT_PATH)
        df['Rerun Attempts'] = count_up_rerun
        df.to_excel(DYNAMIC_EXCEL_REPORT_PATH, index=False)
    else:
        pass


def setRerunCount(testCaseID, rerunCount):
    if ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true":
        try:
            workbook = openpyxl.load_workbook(xl_RerunCountPath)
            sheet = workbook["Rerun At The End"]

            rowNumber = 2
            columnNumber = 1
            sheet.cell(row=rowNumber, column=columnNumber).value = rerunCount
            workbook.save(xl_RerunCountPath)
            workbook.close()
            print(pd.read_excel(xl_RerunCountPath))
            return True
        except:
            return False

    if (ConfigReader.read_config("Validations", "bool_rerun_immediately").lower() == "true" and
            ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "false"):
        try:
            workbook = openpyxl.load_workbook(xl_RerunCountPath)
            sheet = workbook["Rerun Count"]

            rowNumber = ExcelProcessor.getRowNumberFromValue(workbook, sheet, 'Test Case ID',
                                                             testCaseID)
            columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Rerun Count')
            sheet.cell(row=rowNumber, column=columnNumber).value = rerunCount

            workbook.save(xl_RerunCountPath)
            workbook.close()
            return True
        except:
            return False