#!/usr/bin/env python3
"""
Register a generated ReArch test case in DataProvider/TestCasesDetail.xlsx.

Usage:
    python Tools/register_testcase.py <path_to_test_file>

Example:
    python Tools/register_testcase.py TestCases/Functional/UI/ReArch/test_UI_ReArch_PM_BharatQR_Expiry_HDFC_01.py
"""
import re
import subprocess
import sys
import time
from pathlib import Path

import openpyxl


def _close_file_owner(xlsx_path: Path):
    """If a process has xlsx_path open, kill it and wait for the lock file to clear."""
    try:
        result = subprocess.run(
            ["lsof", str(xlsx_path)],
            capture_output=True, text=True
        )
        pids = [
            line.split()[1]
            for line in result.stdout.splitlines()[1:]
            if line.strip()
        ]
        if not pids:
            return
        for pid in set(pids):
            print(f"[INFO] Closing process {pid} that has {xlsx_path.name} open...")
            subprocess.run(["kill", pid], capture_output=True)
        # Wait for lock file to disappear (up to 5 s)
        lock_file = xlsx_path.parent / f".~lock.{xlsx_path.name}#"
        for _ in range(10):
            time.sleep(0.5)
            if not lock_file.exists():
                break
    except Exception as e:
        print(f"[WARN] Could not auto-close file owner: {e}")

XLSX_PATH = Path(__file__).parent.parent / "DataProvider" / "TestCasesDetail.xlsx"

COLUMNS = ["Test Case ID", "Sub Feature Code", "File Name", "Directory Name", "Execute"]


def parse_file(path: Path):
    """Extract xlsx row values from a generated test file."""
    text = path.read_text(encoding="utf-8")

    # Test Case ID — function name
    m = re.search(r'def (test_\w+)\(\):', text)
    if not m:
        raise ValueError(f"No test function found in {path}")
    tc_id = m.group(1)

    # Sub Feature Code — from docstring
    m = re.search(r'Sub Feature Code:\s*(.+)', text)
    sub_feature_code = m.group(1).strip() if m else ""

    # File Name — relative to TestCases/, no .py extension
    abs_path = path.resolve()
    parts = abs_path.parts
    try:
        tc_idx = next(i for i, p in enumerate(parts) if p == "TestCases")
    except StopIteration:
        raise ValueError(f"'TestCases' directory not found in path: {path}")

    rel_dir = "/".join(parts[tc_idx + 1:-1])       # e.g. Functional/UI/ReArch
    file_name = rel_dir + "/" + abs_path.stem       # e.g. Functional/UI/ReArch/test_UI_...
    dir_name = "/".join(rel_dir.split("/")[1:])     # e.g. UI/ReArch (strip "Functional/")

    return tc_id, sub_feature_code, file_name, dir_name


def col_map(ws):
    """Return {header_name: column_index} for named columns in row 1."""
    return {
        ws.cell(1, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def find_or_create_sheet(wb):
    """
    Locate the sheet that holds ReArch tests, in order:
      1. Any sheet whose name contains 'rearch' (case-insensitive)
      2. Any sheet with a File Name cell containing 'Functional/UI/ReArch'
      3. Create a new 'ReArch' sheet with the same headers as the first sheet
    """
    # 1. Name match
    for name in wb.sheetnames:
        if "rearch" in name.lower():
            return wb[name]

    # 2. Content match — look for rows under the File Name column
    for name in wb.sheetnames:
        ws = wb[name]
        cm = col_map(ws)
        fn_col = cm.get("File Name")
        if fn_col is None:
            continue
        for row in ws.iter_rows(min_row=2, min_col=fn_col, max_col=fn_col, values_only=True):
            if row[0] and "Functional/UI/ReArch" in str(row[0]):
                return ws

    # 3. Create new sheet with the same headers as the first sheet
    first_ws = wb[wb.sheetnames[0]]
    headers = [first_ws.cell(1, c).value for c in range(1, first_ws.max_column + 1)]
    new_ws = wb.create_sheet("ReArch")
    new_ws.append(headers)
    print(f"[INFO] Created new sheet 'ReArch' in {XLSX_PATH.name}")
    return new_ws


def register(file_path: str):
    path = Path(file_path)
    if not path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    tc_id, sfc, fn, dn = parse_file(path)

    _close_file_owner(XLSX_PATH)
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
    except PermissionError:
        print(f"[ERROR] {XLSX_PATH.name} is still locked — could not close it automatically.")
        sys.exit(1)

    ws = find_or_create_sheet(wb)
    cm = col_map(ws)

    # Duplicate guard — scan Test Case ID column
    tc_col = cm.get("Test Case ID")
    if tc_col:
        for row in ws.iter_rows(min_row=2, min_col=tc_col, max_col=tc_col, values_only=True):
            if str(row[0]) == tc_id:
                print(f"[SKIP] {tc_id} is already registered in sheet '{ws.title}'")
                return

    # Build new row aligned to existing columns
    new_row = [None] * ws.max_column
    for col_name, val in [
        ("Test Case ID",     tc_id),
        ("Sub Feature Code", sfc),
        ("File Name",        fn),
        ("Directory Name",   dn),
        ("Execute",          True),
    ]:
        if col_name in cm:
            new_row[cm[col_name] - 1] = val

    ws.append(new_row)
    wb.save(XLSX_PATH)
    print(f"[OK] Registered {tc_id} in sheet '{ws.title}' of {XLSX_PATH.name}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    register(sys.argv[1])
