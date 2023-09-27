import datetime
import fcntl
import json
import os
import time
from random import randint
import shutil
from selenium.webdriver.chrome import webdriver
from allure_commons.types import AttachmentType
import allure
import paramiko
from playwright.sync_api import Playwright
from selenium import webdriver
from appium import webdriver as app_webdriver
from datetime import datetime
import chromedriver_autoinstaller
from Utilities import ResourceAssigner, android_utilities
from PageFactory import Base_Actions
from Configuration import TestSuiteSetup
from Utilities import ConfigReader, DirectoryCreator, LogProcessor, Rerun
from pathlib import Path
import openpyxl
import pytest
from termcolor import colored
from DataProvider import GlobalVariables
from Utilities import ExcelProcessor
from Utilities.ReportProcessor import revert_global_variables_to_default, \
    updateExcel_With_Deselect_And_Broken, updateExcel_With_RerunAttempts, updateExcel_With_Category_And_Subcategory
from Utilities.time_calculator import EzeAutoTimeCalculator

tests_count = 0
passed1 = 0
failed1 = 0
skipped1 = 0
changeHTMLVal = 0
rerunCount = 0
currentTestCase = ""
list_deselected_testcases=[]
now = datetime.now()
starting_time = now.strftime("%H:%M:%S")

EXCEL_reportFilePath = DirectoryCreator.getDirectoryPath("ExcelReport")+"/Report.xlsx"

router_ip = Base_Actions.get_environment("str_exe_env_ip")  # dev11
router_username = Base_Actions.get_environment("str_ssh_username")
router_port = Base_Actions.get_environment("int_exe_env_port")
key_filename = Base_Actions.get_environment("str_ssh_key_filename")
ssh = paramiko.SSHClient()


@pytest.mark.hookwrapper
@pytest.hookimpl(hookwrapper=True, tryfirst=True)
def pytest_runtest_makereport(item, call):
    global passed1
    global failed1
    global skipped1
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    setattr(report, "duration_formatter", "%S")  # "%H:%M:%S.%f"
    setattr(item, "rep_" + report.when, report)
    report.description = str(item.function.__doc__)

    if report.outcome == "passed" and report.when == "call":
        passed1 = passed1 + 1

    if report.outcome == "failed" and report.when == "call":
        failed1 = failed1 + 1

    if report.outcome == "skipped":
        skipped1 = skipped1 + 1

    return report


def convert_DF_To_Excel():
    i = randint(1, 10)
    file_name = 'MarksData.xlsx'

    for ind in GlobalVariables.df_testCasesDetail.index:
        testcase_Execution_value = (GlobalVariables.df_testCasesDetail['TC Execution'][ind])

        if testcase_Execution_value != testcase_Execution_value:  # the value is NaN
            GlobalVariables.df_testCasesDetail.at[ind, 'TC Execution'] = "Deselected"
            GlobalVariables.df_testCasesDetail.at[ind, 'API Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'DB Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'Portal Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'App Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'UI Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'ChargeSlip Val'] = "N/A"
            GlobalVariables.df_testCasesDetail.at[ind, 'Execution Time (sec)'] = 0.0
            GlobalVariables.df_testCasesDetail.at[ind, 'Validation Time (sec)'] = 0.0
            GlobalVariables.df_testCasesDetail.at[ind, 'Log Coll Time (sec)'] = 0.0
            GlobalVariables.df_testCasesDetail.at[ind, 'Total Time (sec)'] = 0.0
    # Converting to excel from DF
    GlobalVariables.df_testCasesDetail.to_excel(file_name)


def updatingHighLevelReportAfterEachTCS():
    timer = 0
    while timer < 10:
        try:
            with open(EXCEL_reportFilePath, 'a') as locked_file:
                fcntl.flock(locked_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
                workbook = openpyxl.load_workbook(EXCEL_reportFilePath)
                sheet = workbook["Sheet1"]
                GlobalVariables.EXCEL_testCaseName = os.environ.get('PYTEST_CURRENT_TEST').replace(" (teardown)", '').split('::')[1]
                print("Testcase name", GlobalVariables.EXCEL_testCaseName)
                Overall_Status = 'Broken'
                if GlobalVariables.EXCEL_TC_Execution == 'Pass':

                    # Pass or N/A
                    if GlobalVariables.str_api_val_result != 'Fail' \
                        and GlobalVariables.str_app_val_result != 'Fail' \
                            and GlobalVariables.str_db_val_result != 'Fail' \
                                and GlobalVariables.str_portal_val_result != 'Fail' \
                                    and GlobalVariables.str_ui_val_result != 'Fail'\
                                        and GlobalVariables.str_chargeslip_val_result != "Fail":
                        Overall_Status = 'Pass'

                    elif GlobalVariables.str_api_val_result == 'Fail' \
                        or GlobalVariables.str_app_val_result == 'Fail' \
                            or GlobalVariables.str_db_val_result == 'Fail' \
                                or GlobalVariables.str_portal_val_result == 'Fail' \
                                    or GlobalVariables.str_ui_val_result == 'Fail'\
                                        or GlobalVariables.str_chargeslip_val_result == 'Fail':
                        Overall_Status = 'Fail'
                elif GlobalVariables.EXCEL_TC_Execution == 'Fail':
                    Overall_Status = 'Fail'
                    set_dne_status()
                rowNumber = ExcelProcessor.getRowNumberFromValue(workbook, sheet, 'Test Case ID',
                                                                 GlobalVariables.EXCEL_testCaseName)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'File Name')
                GlobalVariables.EXCEL_testCaseFileName = sheet.cell(row=rowNumber, column=columnNumber).value

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'OverAll Results')
                sheet.cell(row=rowNumber, column=columnNumber).value = Overall_Status

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'TC Execution')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_TC_Execution

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'API Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_api_val_result
                print("API Val: ", GlobalVariables.str_api_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'DB Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_db_val_result
                print("DB Val: ", GlobalVariables.str_db_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Portal Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_portal_val_result
                print("Portal Val: ", GlobalVariables.str_portal_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'App Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_app_val_result
                print("App Val: ", GlobalVariables.str_app_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'UI Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_ui_val_result
                print("UI Val: ", GlobalVariables.str_ui_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'ChargeSlip Val')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_chargeslip_val_result
                print("ChargeSlip Val: ", GlobalVariables.str_chargeslip_val_result)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'API Resp Code')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_api_response_code
                print("API Response Code: ", GlobalVariables.str_api_response_code)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'API Resp Time(sec)')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_api_response_time
                print("API Response Time: ", GlobalVariables.str_api_response_time)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'API Resp Size(kb)')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_api_response_size
                print("API Response Size: ", GlobalVariables.str_api_response_size)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Device Model')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_device_model
                print("Device Model: ", GlobalVariables.str_device_model)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Firmware Version')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_firmware_version
                print("Firmware Version: ", GlobalVariables.str_firmware_version)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'MPOS Version')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_MPOS_version
                print("MPOS Version: ", GlobalVariables.str_MPOS_version)

                columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'SA Version')
                sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.str_SA_version
                print("SA Version: ", GlobalVariables.str_SA_version)

                if (ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true" and ConfigReader.read_config(
                        "Validations", "bool_rerun_immediately").lower() == "false") \
                        or (
                        ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true" and ConfigReader.read_config(
                    "Validations", "bool_rerun_immediately").lower() == "true"):
                    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Rerun Attempts')
                    if sheet.cell(row=rowNumber, column=columnNumber).value is None or sheet.cell(row=rowNumber, column=columnNumber).value == 'N/A':
                        sheet.cell(row=rowNumber, column=columnNumber).value =  0
                    else:

                        currentRetryCountsheet = sheet.cell(row=rowNumber, column=columnNumber).value
                        sheet.cell(row=rowNumber, column=columnNumber).value = currentRetryCountsheet + 1
                # =====================================================================================

                workbook.save(EXCEL_reportFilePath)
                workbook.close()
                fcntl.flock(locked_file, fcntl.LOCK_UN)
                break
        except Exception as e:
            print(f"Unable to update the tc result to report due to error {str(e)}. Retrying..")
            timer += 1
            time.sleep(1)


def set_dne_status():
    if ConfigReader.read_config("Validations", "api_validation").lower() == "true" and 'apiVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_api_val_result = 'DNE'
    if ConfigReader.read_config("Validations", "db_validation").lower() == "true" and 'dbVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_db_val_result = 'DNE'
    if ConfigReader.read_config("Validations", "portal_validation").lower() == "true" and 'portalVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_portal_val_result = 'DNE'
    if ConfigReader.read_config("Validations", "app_validation").lower() == "true" and 'appVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_app_val_result = 'DNE'
    if ConfigReader.read_config("Validations", "ui_validation").lower() == "true" and 'uiVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_ui_val_result = 'DNE'
    if ConfigReader.read_config("Validations", "charge_slip_validation").lower() == "true" and 'chargeSlipVal' in GlobalVariables.tc_markers:
        GlobalVariables.str_chargeslip_val_result = 'DNE'


@pytest.fixture(scope="function")  # Executing once before every testcases
def method_setup(request, playwright: Playwright):
    if GlobalVariables.time_calc is not None:
        GlobalVariables.time_calc.setup.resume()
        print(colored("Setup Timer resumed in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        GlobalVariables.time_calc = EzeAutoTimeCalculator()
        print(colored("Intializaed EzeAutoTimeCalculator in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))
        GlobalVariables.time_calc.setup.start()
        print(colored("Setup Timer started in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    GlobalVariables.tc_markers = [m.name for m in request.node.iter_markers()]
    GlobalVariables.time_calc.setup.pause()
    print(colored("Setup Timer paused in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    GlobalVariables.time_calc.log_collection.start()
    print(colored("Log Collection Timer started in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    print("Function setup level")
    GlobalVariables.LogCollTime = LogProcessor.startLineNoOfServerLogFile()

    GlobalVariables.time_calc.log_collection.pause()
    print(colored("Log Collection Timer paused in method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    GlobalVariables.play_wright = playwright
    # browser = GlobalVariables.play_wright.chromium.launch(headless=False)
    # GlobalVariables.context = browser.new_context(viewport={'width': 1920, 'height': 1080})
    # context_1 = browser.new_context(
    #     record_video_dir="videos/",
    #     record_video_size={"width": 640, "height": 480}
    # )

    # GlobalVariables.portal_page = context.new_page()
    # GlobalVariables.charge_slip_page = context.new_page()
    # breakpoint()
    print("***************End of setup")

    # Executing once AFTER every testcases
    def fin():
        print()
        GlobalVariables.time_calc.teardown.start()
        print(colored("Teardown Timer started in 'fin' of method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

        print("Function teardown level")
        android_utilities.set_report_variables()
        print(GlobalVariables.portal_page)
        print(GlobalVariables.ui_page)
        print(GlobalVariables.charge_slip_page)
        ss_on_failure(request)
        ss_on_success(request)
        # time.sleep(10)
        # context_1.close()
        # if GlobalVariables.str_chargeslip_val_result == "Pass" and GlobalVariables.charge_slip_page != '' and Base_Actions.is_ss_capture_required(
        #         "bool_capt_ss_pass") == "True":
        #     allure.attach("videos/", name="video", attachment_type=AttachmentType.MP4)
        if GlobalVariables.charge_slip_page != '':
            print("closing chargeslip page")
            GlobalVariables.charge_slip_page.close()
            GlobalVariables.charge_slip_page = ''
        if GlobalVariables.ui_page != '':
            print("closing ui_page")
            GlobalVariables.ui_page.close()
            GlobalVariables.ui_page = ''
        if GlobalVariables.portal_page != '':
            print("closing portal_page")
            GlobalVariables.portal_page.close()
            GlobalVariables.portal_page = ''
        if GlobalVariables.context != '':
            print("closing context")
            GlobalVariables.context.close()
            GlobalVariables.context = ''
        if GlobalVariables.browser != '':
            print("closing browser")
            GlobalVariables.browser.close()
            GlobalVariables.browser = ''
        GlobalVariables.time_calc.teardown.pause()
        print(colored("Teardown Timer paused in 'fin' of method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

        GlobalVariables.time_calc.teardown.resume()
        print(colored("Teardown Timer resume in 'fin' of method_setup fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

        updatingHighLevelReportAfterEachTCS()
        item = request.node
        if item.rep_call.failed:
            log_on_failure(request)

        if (ConfigReader.read_config("Validations", "bool_rerun_immediately").lower() == "true" and
                Rerun.isRerunRequiredImmediately(GlobalVariables.EXCEL_testCaseName) and
                ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "false"):
            rerunCount = Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName)

            if rerunCount >= 0:
                print(str(rerunCount) + " reruns pending for the test case " + GlobalVariables.EXCEL_testCaseName)
                rerunCount -= 1
                Rerun.rerunTestImmediately(GlobalVariables.EXCEL_testCaseName, GlobalVariables.EXCEL_testCaseFileName,
                                           rerunCount, request)
            else:
                print(str(rerunCount) + " reruns pending for the test case " + GlobalVariables.EXCEL_testCaseName)
                print("Rerun skipped.")

        revert_global_variables_to_default()
        GlobalVariables.time_calc.teardown.end()
        print(colored("Teardown Timer ended in 'fin' -> 'method_setup' fixture".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    request.addfinalizer(fin)


def write_TC_Details_To_Dataframe():
    print("################", GlobalVariables.str_portal_val_result)
    TestCaseName = os.environ.get('PYTEST_CURRENT_TEST').replace(" (teardown)", '').replace("test_sample.py::",
                                                                                            '').replace("TestCase/", '')
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'TC Execution'] = GlobalVariables.EXCEL_TC_Execution
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'API Val'] = GlobalVariables.str_api_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'DB Val'] = GlobalVariables.str_db_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'Portal Val'] = GlobalVariables.str_portal_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'App Val'] = GlobalVariables.str_app_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'UI Val'] = GlobalVariables.str_ui_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'ChargeSlip Val'] = GlobalVariables.str_chargeslip_val_result
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'Execution Time (sec)'] = GlobalVariables.EXCEL_Execution_Time
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'Validation Time (sec)'] = GlobalVariables.EXCEL_Val_time
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'Log Coll Time (sec)'] = GlobalVariables.EXCEL_LogCollTime
    GlobalVariables.df_testCasesDetail.at[TestCaseName, 'Total Time (sec)'] = GlobalVariables.EXCEL_Tot_Time


@pytest.fixture(scope="function")
def ui_driver(request):
    if GlobalVariables.time_calc is not None:
        GlobalVariables.time_calc.setup.resume()
        print(colored("Setup Timer resumed in ui_driver function".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        GlobalVariables.time_calc = EzeAutoTimeCalculator()
        print(colored("Intializaed EzeAutoTimeCalculator in ui_driver".center(shutil.get_terminal_size().columns, "="), 'cyan'))
        GlobalVariables.time_calc.setup.start()
        print(colored("Setup Timer started in ui_driver".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    GlobalVariables.portalDriver = chromedriver_autoinstaller.install()
    # Chrome options
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument('--disable-dev-shm-usage')
    # Run chrome
    GlobalVariables.portalDriver = webdriver.Chrome(options=chrome_options)
    GlobalVariables.portalDriver.maximize_window()

    GlobalVariables.time_calc.setup.pause()
    print(colored("Setup Timer paused in ui_driver function".center(shutil.get_terminal_size().columns, "="), 'cyan'))


@pytest.fixture(scope="function")
def appium_driver(request):
    if GlobalVariables.time_calc is not None:
        GlobalVariables.time_calc.setup.resume()
        print(colored("Setup Timer resumed in appium_driver function".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        GlobalVariables.time_calc = EzeAutoTimeCalculator()
        print(colored("Intializaed EzeAutoTimeCalculator in appium_driver".center(shutil.get_terminal_size().columns, "="), 'cyan'))
        GlobalVariables.time_calc.setup.start()
        print(colored("Setup Timer started in appium_driver".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    testcaseid = request.node.name
    deviceDetails = ResourceAssigner.getDeviceFromDB(testcaseid)
    appiumserverDetails = ResourceAssigner.getAppiumServerFromDB(testcaseid)
    print(testcaseid+" will be using the device "+deviceDetails['DeviceId'])
    print(testcaseid + " will be running on the appium server port " + appiumserverDetails['PortNumber'])
    mposApp = ConfigReader.read_config_paths("System","automation_suite_path")+"/App/"+ConfigReader.read_config("Applications","mpos")
    saApp = ConfigReader.read_config_paths("System","automation_suite_path")+"/App/"+ConfigReader.read_config("Applications","SA")
    lst_applications = [mposApp,saApp]
    json_applications = json.dumps(lst_applications)
    desired_cap = {
        "platformName": "Android",
        "deviceName": deviceDetails['DeviceId'],
        "udid": deviceDetails['DeviceId'],
        "otherApps": json_applications,
        "appPackage": "com.ezetap.basicapp",
        "appActivity": "com.ezetap.mposX.activity.SplashActivity",
        "ignoreHiddenApiPolicyError": "true",
        "noReset": "false",
        "autoGrantPermissions": "true",
        "newCommandTimeout":7000,
        "MobileCapabilityType.AUTOMATION_NAME": "AutomationName.ANDROID_UIAUTOMATOR2",
        "MobileCapabilityType.NEW_COMMAND_TIMEOUT":"300"
    }
    print(desired_cap)
    print("appium server url:", 'http://127.0.0.1:' + appiumserverDetails['PortNumber'] + '/wd/hub')
    GlobalVariables.appDriver = app_webdriver.Remote('http://127.0.0.1:' + appiumserverDetails['PortNumber'] + '/wd/hub', desired_cap)
    # GlobalVariables.appDriver.implicitly_wait(30)

    GlobalVariables.time_calc.setup.pause()
    print(colored("Setup Timer paused in appium_driver function".center(shutil.get_terminal_size().columns, "="), 'cyan'))


def pytest_deselected(items):
    print("INSIDE DESELCTED METHOD")
    for item in items:
        print(item.nodeid)
        testcase = str((item.nodeid)).split('::')[1]
        GlobalVariables.list_deselected_testcases.append(testcase)


def isBothRerunNotEnabled():
    if ConfigReader.read_config("Validations","bool_rerun_at_the_end").lower() == "false" and ConfigReader.read_config("Validations","bool_rerun_immediately").lower() == "false":
        return True
    return False


def ss_on_failure(request):
    item = request.node
    if item.rep_call.failed:
        if GlobalVariables.bool_ss_app_val == 'Failed' and GlobalVariables.appDriver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.appDriver.get_screenshot_as_png(), name="app_screen",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.bool_ss_app_val = 'Passed'

        if GlobalVariables.bool_ss_portal_val == 'Failed' and GlobalVariables.portalDriver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.portalDriver.get_screenshot_as_png(), name="portal_page",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.bool_ss_portal_val = 'Passed'

        if GlobalVariables.str_chargeslip_val_result == "Fail" and GlobalVariables.charge_slip_driver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.charge_slip_driver.get_screenshot_as_png(), name="chargeslip",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.str_chargeslip_val_result == "Fail" and GlobalVariables.charge_slip_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.charge_slip_page.screenshot(full_page=True), name="chargeslip",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.EXCEL_TC_Execution == "Fail" and GlobalVariables.ui_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.ui_page.screenshot(full_page=True), name="Ui_browser",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.str_portal_val_result == "Fail" and GlobalVariables.portal_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_fail") == "True":
            allure.attach(GlobalVariables.portal_page.screenshot(full_page=True), name="portal_browser",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.portalDriver != '':
            # GlobalVariables.portalDriver.quit()
            GlobalVariables.portalDriver.close()
            GlobalVariables.portalDriver = ''
            # variables.successApp = False

        if GlobalVariables.appDriver != '':
            GlobalVariables.appDriver.quit()
            GlobalVariables.appDriver = ''

        if GlobalVariables.charge_slip_driver != '':
            GlobalVariables.charge_slip_driver.quit()
            GlobalVariables.charge_slip_driver = ''

        if GlobalVariables.charge_slip_page != '':
            print("closing chargeslip page")
            GlobalVariables.charge_slip_page.close()
            GlobalVariables.charge_slip_page = ''
        if GlobalVariables.ui_page != '':
            print("closing ui_page")
            GlobalVariables.ui_page.close()
            GlobalVariables.ui_page = ''
        if GlobalVariables.portal_page != '':
            print("closing portal_page")
            GlobalVariables.portal_page.close()
            GlobalVariables.portal_page = ''
        if GlobalVariables.context != '':
            print("closing context")
            GlobalVariables.context.close()
            GlobalVariables.context = ''
        if GlobalVariables.browser != '':
            print("closing browser")
            GlobalVariables.browser.close()
            GlobalVariables.browser = ''
        GlobalVariables.bool_ss_portal_val = "N/A"
        GlobalVariables.bool_ss_app_val = "N/A"


def log_on_failure(request):
    if GlobalVariables.time_calc.log_collection.is_started and GlobalVariables.time_calc.log_collection.is_paused:
        GlobalVariables.time_calc.log_collection.resume()
        print(colored("Log Collection Timer resumed in 'log_on_failure' function of conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        GlobalVariables.time_calc.log_collection.start()
        print(colored("Log Collection started in 'log_on_failure' function of conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    item = request.node
    if item.rep_call.failed:
        if Base_Actions.is_log_capture_required("bool_capt_log_fail") == "True":
            if (ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true" or
                    ConfigReader.read_config("Validations", "bool_rerun_immediately").lower() == "true"):
                if Base_Actions.is_log_capture_required("bool_capt_log_one_file") == "True":
                    if Base_Actions.is_log_capture_required("bool_capt_log_each_run") == "True":
                        path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"
                        print("Inside capturing logs all in one file")

                        Path(path).mkdir(parents=True, exist_ok=True)
                        TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())
                        if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_api") == "True":
                            apiLogs = LogProcessor.fetch_API_logs()
                            print("TCIdWithTimeStamp", TCIdWithTimeStamp)
                            rerun_file = Path(path + "/api.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                        if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_portal") == "True":
                            portalLogs = LogProcessor.fetch_portal_logs()
                            rerun_file = Path(path + "/portal.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                        if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_middleware") == "True":
                            mwareLogs = LogProcessor.fetch_middleware_logs()
                            rerun_file = Path(path + "/middleware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                        if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_cnpware") == "True":
                            cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                            rerun_file = Path(path + "/cnpware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                        if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_config") == "True":
                            config_logs = LogProcessor.fetch_config_logs()
                            rerun_file = Path(path + "/config.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                        if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_closedloop") == "True":
                            closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                            rerun_file = Path(path + "/closedloop.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                        if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_q2") == "True":
                            q2_logs = LogProcessor.fetch_q2_logs()
                            rerun_file = Path(path + "/q2.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                        if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_commx") == "True":
                            commx_logs = LogProcessor.fetch_commx_logs()
                            rerun_file = Path(path + "/commx.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commx_logs)

                        if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_ezestore") == "True":
                            ezestore_logs = LogProcessor.fetch_ezestore_logs()
                            rerun_file = Path(path + "/ezestore.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestore_logs)

                        if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_khata") == "True":
                            khata_logs = LogProcessor.fetch_khata_logs()
                            rerun_file = Path(path + "/khata.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                        if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_reward") == "True":
                            reward_logs = LogProcessor.fetch_reward_logs()
                            rerun_file = Path(path + "/reward.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                    else:
                        if Base_Actions.is_log_capture_required("bool_capt_log_last_run") == "True":
                            print("Inside capturing log of last run TCs in one file")

                            if ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true":
                                rerun_count = Rerun.getRerunCount(str(request.node.nodeid))
                            else:
                                print(Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName))
                                rerun_count = Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName)

                            if rerun_count == 0:
                                print("Fetching logs of last run")

                                path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"
                                Path(path).mkdir(parents=True, exist_ok=True)

                                TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                                if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_api") == "True":
                                    apiLogs = LogProcessor.fetch_API_logs()
                                    rerun_file = Path(path + "/api.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                                if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_portal") == "True":
                                    portalLogs = LogProcessor.fetch_portal_logs()
                                    rerun_file = Path(path + "/portal.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                                if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_middleware") == "True":
                                    mwareLogs = LogProcessor.fetch_middleware_logs()
                                    rerun_file = Path(path + "/middleware.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                                if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_cnpware") == "True":
                                    cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                    rerun_file = Path(path + "/cnpware.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                                if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_config") == "True":
                                    config_logs = LogProcessor.fetch_config_logs()
                                    rerun_file = Path(path + "/config.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                                if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_closedloop") == "True":
                                    closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                    rerun_file = Path(path + "/closedloop.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                                if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_q2") == "True":
                                    q2_logs = LogProcessor.fetch_q2_logs()
                                    rerun_file = Path(path + "/q2.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                                if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_commx") == "True":
                                    commxLogs = LogProcessor.fetch_commx_logs()
                                    rerun_file = Path(path + "/commx.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                                if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_ezestore") == "True":
                                    ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                                    rerun_file = Path(path + "/ezestore.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                                if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_khata") == "True":
                                    khata_logs = LogProcessor.fetch_khata_logs()
                                    rerun_file = Path(path + "/khata.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                                if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_reward") == "True":
                                    reward_logs = LogProcessor.fetch_reward_logs()
                                    rerun_file = Path(path + "/reward.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                            else:
                                print("This is not last run")
                        else:
                            print("Both last_run and all_run are disabled")
                else:
                    if Base_Actions.is_log_capture_required("bool_capt_log_different_files") == "True":
                        if Base_Actions.is_log_capture_required("bool_capt_log_each_run") == "True":
                            print("bool_capt_log_each_run")
                            i = int(ConfigReader.read_config("Validations", "int_rerun_count"))
                            j = 0
                            print("============ i = ",i)
                            print("============ j = ",0)
                            TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                            print("TCIdWithTimeStamp : ",TCIdWithTimeStamp)

                            testCaseID = str(item.nodeid).split('/')
                            finalTestCaseID = testCaseID[len(testCaseID) - 1]
                            logFileName = str(finalTestCaseID).split('::')[1]

                            while i >= 0 and j <= int(ConfigReader.read_config("Validations", "int_rerun_count")):
                                if Rerun.getRerunCount(str(item.nodeid).split('::')[1]) == i:

                                    path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName
                                    Path(path).mkdir(parents=True, exist_ok=True)

                                    if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_api") == "True":
                                        apiLogs = LogProcessor.fetch_API_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/api.log")
                                        else:
                                            rerun_file = Path(path + "/api_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                                    if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_portal") == "True":
                                        portalLogs = LogProcessor.fetch_portal_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/portal.log")
                                        else:
                                            rerun_file = Path(path + "/portal_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                                    if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_middleware") == "True":
                                        mwareLogs = LogProcessor.fetch_middleware_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/middleware.log")
                                        else:
                                            rerun_file = Path(path + "/middleware_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                                    if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_cnpware") == "True":
                                        cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/cnpware.log")
                                        else:
                                            rerun_file = Path(path + "/cnpware_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                                    if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_config") == "True":
                                        config_logs = LogProcessor.fetch_config_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/config.log")
                                        else:
                                            rerun_file = Path(path + "/config_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                                    if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_closedloop") == "True":
                                        closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/closedloop.log")
                                        else:
                                            rerun_file = Path(path + "/closedloop_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                                    if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_q2") == "True":
                                        q2_logs = LogProcessor.fetch_q2_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/q2.log")
                                        else:
                                            rerun_file = Path(path + "/q2_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                                    if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_commx") == "True":
                                        commxLogs = LogProcessor.fetch_commx_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/commx.log")
                                        else:
                                            rerun_file = Path(path + "/commx_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                                    if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_ezestore") == "True":
                                        ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/ezestore.log")
                                        else:
                                            rerun_file = Path(path + "/ezestore_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                                    if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_khata") == "True":
                                        khata_logs = LogProcessor.fetch_khata_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/khata.log")
                                        else:
                                            rerun_file = Path(path + "/khata_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                                    if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_reward") == "True":
                                        reward_logs = LogProcessor.fetch_reward_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/reward.log")
                                        else:
                                            rerun_file = Path(path + "/reward_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)

                                i -= 1
                                j += 1
                        else:
                            print("bool_capt_log_each_run == false")
                            if Base_Actions.is_log_capture_required("bool_capt_log_last_run") == "True":
                                print("bool_capt_log_last_run == true")
                                if ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "true":
                                    rerun_count = Rerun.getRerunCount(str(request.node.nodeid))
                                else:
                                    print(Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName))
                                    rerun_count = Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName)
                                if rerun_count == 0:
                                    print("Fetching logs of last run")
                                    testCaseID = str(item.nodeid).split('/')
                                    finalTestCaseID = testCaseID[len(testCaseID) - 1]
                                    logFileName = str(finalTestCaseID).split('::')[1]

                                    path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName
                                    Path(path).mkdir(parents=True, exist_ok=True)

                                    TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                                    if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_api") == "True":
                                        apiLogs = LogProcessor.fetch_API_logs()
                                        rerun_file = Path(path + "/api.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                                    if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_portal") == "True":
                                        portalLogs = LogProcessor.fetch_portal_logs()
                                        rerun_file = Path(path + "/portal.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                                    if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_middleware") == "True":
                                        mwareLogs = LogProcessor.fetch_middleware_logs()
                                        rerun_file = Path(path + "/middleware.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                                    if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_cnpware") == "True":
                                        cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                        rerun_file = Path(path + "/cnpware.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                                    if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_config") == "True":
                                        config_logs = LogProcessor.fetch_config_logs()
                                        rerun_file = Path(path + "/config.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                                    if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_closedloop") == "True":
                                        closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                        rerun_file = Path(path + "/closedloop.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                                    if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_q2") == "True":
                                        q2_logs = LogProcessor.fetch_q2_logs()
                                        rerun_file = Path(path + "/q2.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                                    if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_commx") == "True":
                                        commxLogs = LogProcessor.fetch_commx_logs()
                                        rerun_file = Path(path + "/commx.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                                    if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_ezestore") == "True":
                                        ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                                        rerun_file = Path(path + "/ezestore.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                                    if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_khata") == "True":
                                        khata_logs = LogProcessor.fetch_khata_logs()
                                        rerun_file = Path(path + "/khata.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                                    if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_reward") == "True":
                                        reward_logs = LogProcessor.fetch_reward_logs()
                                        rerun_file = Path(path + "/reward.log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                                else:
                                    print("This is not last run")
                            else:
                                print("bool_capt_log_last_run == false")
                                print("Both last_run and each_run are disabled")
                    else:
                        print("Both Same file and different file are disabled")
            else:
                print("Rerun is disabled")
                if Base_Actions.is_log_capture_required("bool_capt_log_one_file") == "True":
                    path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"
                    print("Inside capturing logs all in one file")
                    print("Fetching Logs 89th time")

                    Path(path).mkdir(parents=True, exist_ok=True)
                    TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())
                    if GlobalVariables.api_logs and Base_Actions.is_log_capture_required("bool_capt_log_api") == "True":
                        apiLogs = LogProcessor.fetch_API_logs()
                        print("TCIdWithTimeStamp", TCIdWithTimeStamp)
                        rerun_file = Path(path + "/api.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                    if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_portal") == "True":
                        portalLogs = LogProcessor.fetch_portal_logs()
                        rerun_file = Path(path + "/portal.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                    if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_middleware") == "True":
                        mwareLogs = LogProcessor.fetch_middleware_logs()
                        rerun_file = Path(path + "/middleware.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                    if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_cnpware") == "True":
                        cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                        rerun_file = Path(path + "/cnpware.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                    if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_config") == "True":
                        config_logs = LogProcessor.fetch_config_logs()
                        rerun_file = Path(path + "/config.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                    if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_closedloop") == "True":
                        closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                        rerun_file = Path(path + "/closedloop.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                    if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required("bool_capt_log_q2") == "True":
                        q2_logs = LogProcessor.fetch_q2_logs()
                        rerun_file = Path(path + "/q2.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                    if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_commx") == "True":
                        commx_logs = LogProcessor.fetch_commx_logs()
                        rerun_file = Path(path + "/commx.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commx_logs)

                    if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_ezestore") == "True":
                        ezestore_logs = LogProcessor.fetch_ezestore_logs()
                        rerun_file = Path(path + "/ezestore.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestore_logs)

                    if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_khata") == "True":
                        khata_logs = LogProcessor.fetch_khata_logs()
                        rerun_file = Path(path + "/khata.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                    if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_reward") == "True":
                        reward_logs = LogProcessor.fetch_reward_logs()
                        rerun_file = Path(path + "/reward.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                else:
                    if Base_Actions.is_log_capture_required("bool_capt_log_different_files") == "True":
                        testCaseID = str(item.nodeid).split('/')
                        finalTestCaseID = testCaseID[len(testCaseID) - 1]
                        logFileName = str(finalTestCaseID).split('::')[1]

                        path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName
                        Path(path).mkdir(parents=True, exist_ok=True)

                        TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                        if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_api") == "True":
                            apiLogs = LogProcessor.fetch_API_logs()
                            rerun_file = Path(path + "/api.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                        if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_portal") == "True":
                            portalLogs = LogProcessor.fetch_portal_logs()
                            rerun_file = Path(path + "/portal.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                        if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_middleware") == "True":
                            mwareLogs = LogProcessor.fetch_middleware_logs()
                            rerun_file = Path(path + "/middleware.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                        if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_cnpware") == "True":
                            cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                            rerun_file = Path(path + "/cnpware.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                        if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_config") == "True":
                            config_logs = LogProcessor.fetch_config_logs()
                            rerun_file = Path(path + "/config.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                        if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_closedloop") == "True":
                            closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                            rerun_file = Path(path + "/closedloop.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                        if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_q2") == "True":
                            q2_logs = LogProcessor.fetch_q2_logs()
                            rerun_file = Path(path + "/q2.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                        if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_commx") == "True":
                            commxLogs = LogProcessor.fetch_commx_logs()
                            rerun_file = Path(path + "/commx.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                        if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_ezestore") == "True":
                            ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                            rerun_file = Path(path + "/ezestore.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                        if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_khata") == "True":
                            khata_logs = LogProcessor.fetch_khata_logs()
                            rerun_file = Path(path + "/khata.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                        if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_reward") == "True":
                            reward_logs = LogProcessor.fetch_reward_logs()
                            rerun_file = Path(path + "/reward.log")
                            open(rerun_file, 'w').close()
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                    else:
                        print("Both bool_capt_log_one_file and bool_capt_log_different_files are disabled")

    GlobalVariables.time_calc.log_collection.pause()
    print(colored("Log Collection Timer paused in 'log on failure' function in conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))


def ss_on_success(request):
    item = request.node
    if item.rep_call.passed:
        if GlobalVariables.bool_ss_app_val == 'Passed' and GlobalVariables.appDriver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            allure.attach(GlobalVariables.appDriver.get_screenshot_as_png(), name="app_screen",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.bool_ss_app_val = 'Passed'

        if GlobalVariables.bool_ss_portal_val == 'Passed' and GlobalVariables.portalDriver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            allure.attach(GlobalVariables.portalDriver.get_screenshot_as_png(), name="portal_page",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.bool_ss_portal_val = 'Passed'

        if GlobalVariables.str_chargeslip_val_result == "Pass" and GlobalVariables.charge_slip_driver != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            allure.attach(GlobalVariables.charge_slip_driver.get_screenshot_as_png(), name="chargeslip",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.str_chargeslip_val_result == "Pass" and GlobalVariables.charge_slip_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            # GlobalVariables.charge_slip_page.wait_for_load_state('networkidle', timeout=50000)
            allure.attach(GlobalVariables.charge_slip_page.screenshot(full_page=True), name="chargeslip",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.EXCEL_TC_Execution == "Pass" and GlobalVariables.ui_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            # if GlobalVariables.ui_page.wait_for_load_state('networkidle', timeout=50000):
            allure.attach(GlobalVariables.ui_page.screenshot(full_page=True), name="ui_browser",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.str_portal_val_result == "Pass" and GlobalVariables.portal_page != '' and Base_Actions.is_ss_capture_required(
                "bool_capt_ss_pass") == "True":
            # GlobalVariables.portal_page.wait_for_load_state('networkidle', timeout=50000)
            allure.attach(GlobalVariables.portal_page.screenshot(full_page=True), name="portal_browser",
                          attachment_type=AttachmentType.PNG)
        if GlobalVariables.portalDriver != '':
            GlobalVariables.portalDriver.close()
            GlobalVariables.portalDriver = ''

        if GlobalVariables.appDriver != '':
            GlobalVariables.appDriver.quit()
            GlobalVariables.appDriver = ''

        if GlobalVariables.charge_slip_driver != '':
            GlobalVariables.charge_slip_driver.quit()
            GlobalVariables.charge_slip_driver = ''

        if GlobalVariables.charge_slip_page != '':
            print("closing chargeslip page")
            GlobalVariables.charge_slip_page.close()
            GlobalVariables.charge_slip_page = ''
        if GlobalVariables.ui_page != '':
            print("closing ui_page")
            GlobalVariables.ui_page.close()
            GlobalVariables.ui_page = ''
        if GlobalVariables.portal_page != '':
            print("closing portal_page")
            GlobalVariables.portal_page.close()
            GlobalVariables.portal_page = ''
        if GlobalVariables.context != '':
            print("closing context")
            GlobalVariables.context.close()
            GlobalVariables.context = ''
        if GlobalVariables.browser != '':
            print("closing browser")
            GlobalVariables.browser.close()
            GlobalVariables.browser = ''

        GlobalVariables.bool_ss_portal_val = "N/A"
        GlobalVariables.bool_ss_app_val = "N/A"


@pytest.fixture(scope='function')
def log_on_success(request):
    yield

    if GlobalVariables.time_calc.log_collection.is_started and GlobalVariables.time_calc.log_collection.is_paused:
        GlobalVariables.time_calc.log_collection.resume()
        print(colored("Log Collection resumed in 'log_on_success' function of conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))
    else:
        GlobalVariables.time_calc.log_collection.start()
        print(colored("Log Collection started in 'log_on_success' function of conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    item = request.node

    if item.rep_call.passed:
        if Base_Actions.is_log_capture_required("bool_capt_log_pass") == "True":
            if ConfigReader.read_config("Validations",
                                        "bool_rerun_at_the_end").lower() == "true" or ConfigReader.read_config("Validations","bool_rerun_immediately").lower() == "true":
                if Base_Actions.is_log_capture_required("bool_capt_log_one_file") == "True":
                    if Base_Actions.is_log_capture_required("bool_capt_log_each_run") == "True":
                        path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"
                        print("Inside capturing logs all in one file")

                        Path(path).mkdir(parents=True, exist_ok=True)
                        TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())
                        if GlobalVariables.api_logs and Base_Actions.is_log_capture_required("bool_capt_log_api") == "True":
                            apiLogs = LogProcessor.fetch_API_logs()
                            print("TCIdWithTimeStamp", TCIdWithTimeStamp)
                            rerun_file = Path(path + "/api.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                        if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_portal") == "True":
                            portalLogs = LogProcessor.fetch_portal_logs()
                            rerun_file = Path(path + "/portal.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                        if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_middleware") == "True":
                            mwareLogs = LogProcessor.fetch_middleware_logs()
                            rerun_file = Path(path + "/middleware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                        if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_cnpware") == "True":
                            cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                            rerun_file = Path(path + "/cnpware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                        if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_config") == "True":
                            config_logs = LogProcessor.fetch_config_logs()
                            rerun_file = Path(path + "/config.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                        if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_closedloop") == "True":
                            closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                            rerun_file = Path(path + "/closedloop.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                        if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_q2") == "True":
                            q2_logs = LogProcessor.fetch_q2_logs()
                            rerun_file = Path(path + "/q2.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                        if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_commx") == "True":
                            commx_logs = LogProcessor.fetch_commx_logs()
                            rerun_file = Path(path + "/commx.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commx_logs)

                        if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_ezestore") == "True":
                            ezestore_logs = LogProcessor.fetch_ezestore_logs()
                            rerun_file = Path(path + "/ezestore.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestore_logs)

                        if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_khata") == "True":
                            khata_logs = LogProcessor.fetch_khata_logs()
                            rerun_file = Path(path + "/khata.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                        if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_reward") == "True":
                            reward_logs = LogProcessor.fetch_reward_logs()
                            rerun_file = Path(path + "/reward.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                    else:
                        if Base_Actions.is_log_capture_required("bool_capt_log_last_run") == "True":
                            print("capt_log_last_run == true")
                            path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"

                            Path(path).mkdir(parents=True, exist_ok=True)
                            TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())
                            if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_api") == "True":
                                apiLogs = LogProcessor.fetch_API_logs()
                                rerun_file = Path(path + "/api.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                            if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_portal") == "True":
                                portalLogs = LogProcessor.fetch_portal_logs()
                                rerun_file = Path(path + "/portal.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                            if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_middleware") == "True":
                                mwareLogs = LogProcessor.fetch_middleware_logs()
                                rerun_file = Path(path + "/middleware.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                            if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_cnpware") == "True":
                                cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                rerun_file = Path(path + "/cnpware.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                            if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_config") == "True":
                                config_logs = LogProcessor.fetch_config_logs()
                                rerun_file = Path(path + "/config.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                            if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_closedloop") == "True":
                                closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                rerun_file = Path(path + "/closedloop.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                            if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_q2") == "True":
                                q2_logs = LogProcessor.fetch_q2_logs()
                                rerun_file = Path(path + "/q2.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                            if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_commx") == "True":
                                commx_logs = LogProcessor.fetch_commx_logs()
                                rerun_file = Path(path + "/commx.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commx_logs)

                            if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_ezestore") == "True":
                                ezestore_logs = LogProcessor.fetch_ezestore_logs()
                                rerun_file = Path(path + "/ezestore.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestore_logs)

                            if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_khata") == "True":
                                khata_logs = LogProcessor.fetch_khata_logs()
                                rerun_file = Path(path + "/khata.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                            if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                    "bool_capt_log_reward") == "True":
                                reward_logs = LogProcessor.fetch_reward_logs()
                                rerun_file = Path(path + "/reward.log")
                                LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                        else:
                            print("Both last_run and all_run are disabled")
                else:
                    if Base_Actions.is_log_capture_required("bool_capt_log_different_files") == "True":
                        if Base_Actions.is_log_capture_required("bool_capt_log_each_run") == "True":
                            i = int(ConfigReader.read_config("Validations", "int_rerun_count"))
                            j = 0

                            TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                            testCaseID = str(item.nodeid).split('/')
                            finalTestCaseID = testCaseID[len(testCaseID) - 1]
                            logFileName = str(finalTestCaseID).split('::')[1]

                            while i >= 0 and j <= int(ConfigReader.read_config("Validations", "int_rerun_count")):
                                if Rerun.getRerunCount(str(item.nodeid).split('::')[1]) == i:
                                    path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName

                                    Path(path).mkdir(parents=True, exist_ok=True)

                                    if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_api") == "True":
                                        apiLogs = LogProcessor.fetch_API_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/api.log")
                                        else:
                                            rerun_file = Path(path + "/api_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                                    if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_portal") == "True":
                                        portalLogs = LogProcessor.fetch_portal_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/portal.log")
                                        else:
                                            rerun_file = Path(path + "/portal_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                                    if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_middleware") == "True":
                                        mwareLogs = LogProcessor.fetch_middleware_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/middleware.log")
                                        else:
                                            rerun_file = Path(path + "/middleware_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                                    if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_cnpware") == "True":
                                        cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/cnpware.log")
                                        else:
                                            rerun_file = Path(path + "/cnpware_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                                    if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_config") == "True":
                                        config_logs = LogProcessor.fetch_config_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/config.log")
                                        else:
                                            rerun_file = Path(path + "/config_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                                    if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_closedloop") == "True":
                                        closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/closedloop.log")
                                        else:
                                            rerun_file = Path(path + "/closedloop_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                                    if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_q2") == "True":
                                        q2_logs = LogProcessor.fetch_q2_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/q2.log")
                                        else:
                                            rerun_file = Path(path + "/q2_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                                    if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_commx") == "True":
                                        commxLogs = LogProcessor.fetch_commx_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/commx.log")
                                        else:
                                            rerun_file = Path(path + "/commx_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                                    if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_ezestore") == "True":
                                        ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/ezestore.log")
                                        else:
                                            rerun_file = Path(path + "/ezestore_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                                    if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_khata") == "True":
                                        khata_logs = LogProcessor.fetch_khata_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/khata.log")
                                        else:
                                            rerun_file = Path(path + "/khata_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                                    if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                            "bool_capt_log_reward") == "True":
                                        reward_logs = LogProcessor.fetch_reward_logs()
                                        if j == 0:
                                            rerun_file = Path(path + "/reward.log")
                                        else:
                                            rerun_file = Path(path + "/reward_Rerun_" + str(j) + ".log")
                                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)

                                i -= 1
                                j += 1
                        else:
                            if Base_Actions.is_log_capture_required("bool_capt_log_last_run") == "True":
                                print("bool_capt_log_last_run == true")
                                testCaseID = str(item.nodeid).split('/')
                                finalTestCaseID = testCaseID[len(testCaseID) - 1]
                                logFileName = str(finalTestCaseID).split('::')[1]

                                path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName
                                Path(path).mkdir(parents=True, exist_ok=True)

                                TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                                if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_api") == "True":
                                    apiLogs = LogProcessor.fetch_API_logs()
                                    rerun_file = Path(path + "/api.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                                if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_portal") == "True":
                                    portalLogs = LogProcessor.fetch_portal_logs()
                                    rerun_file = Path(path + "/portal.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                                if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_middleware") == "True":
                                    mwareLogs = LogProcessor.fetch_middleware_logs()
                                    rerun_file = Path(path + "/middleware.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                                if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_cnpware") == "True":
                                    cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                                    rerun_file = Path(path + "/cnpware.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                                if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_config") == "True":
                                    config_logs = LogProcessor.fetch_config_logs()
                                    rerun_file = Path(path + "/config.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                                if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_closedloop") == "True":
                                    closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                                    rerun_file = Path(path + "/closedloop.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                                if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_q2") == "True":
                                    q2_logs = LogProcessor.fetch_q2_logs()
                                    rerun_file = Path(path + "/q2.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                                if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_commx") == "True":
                                    commxLogs = LogProcessor.fetch_commx_logs()
                                    rerun_file = Path(path + "/commx.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                                if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_ezestore") == "True":
                                    ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                                    rerun_file = Path(path + "/ezestore.log")
                                    open(rerun_file, 'w').close()
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                                if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_khata") == "True":
                                    khata_logs = LogProcessor.fetch_khata_logs()
                                    rerun_file = Path(path + "/khata.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                                if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                        "bool_capt_log_reward") == "True":
                                    reward_logs = LogProcessor.fetch_reward_logs()
                                    rerun_file = Path(path + "/reward.log")
                                    LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                            else:
                                print("Both last_run and each_run are disabled")
                    else:
                        print("Both Same file and different file are disabled")

            else:
                print("Rerun is disabled SUCCESS")
                if Base_Actions.is_log_capture_required("bool_capt_log_one_file") == "True":
                    path = DirectoryCreator.getDirectoryPath("ServerLog") + "/"
                    print("Inside capturing logs all in one file")

                    Path(path).mkdir(parents=True, exist_ok=True)
                    TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())
                    if GlobalVariables.api_logs and Base_Actions.is_log_capture_required("bool_capt_log_api") == "True":
                        apiLogs = LogProcessor.fetch_API_logs()
                        print("TCIdWithTimeStamp", TCIdWithTimeStamp)
                        rerun_file = Path(path + "/api.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                    if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_portal") == "True":
                        portalLogs = LogProcessor.fetch_portal_logs()
                        rerun_file = Path(path + "/portal.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                    if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_middleware") == "True":
                        mwareLogs = LogProcessor.fetch_middleware_logs()
                        rerun_file = Path(path + "/middleware.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                    if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_cnpware") == "True":
                        cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                        rerun_file = Path(path + "/cnpware.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                    if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_config") == "True":
                        config_logs = LogProcessor.fetch_config_logs()
                        rerun_file = Path(path + "/config.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                    if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_closedloop") == "True":
                        closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                        rerun_file = Path(path + "/closedloop.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                    if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required("bool_capt_log_q2") == "True":
                        q2_logs = LogProcessor.fetch_q2_logs()
                        rerun_file = Path(path + "/q2.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                    if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_commx") == "True":
                        commx_logs = LogProcessor.fetch_commx_logs()
                        rerun_file = Path(path + "/commx.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commx_logs)

                    if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_ezestore") == "True":
                        ezestore_logs = LogProcessor.fetch_ezestore_logs()
                        rerun_file = Path(path + "/ezestore.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestore_logs)

                    if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_khata") == "True":
                        khata_logs = LogProcessor.fetch_khata_logs()
                        rerun_file = Path(path + "/khata.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                    if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                            "bool_capt_log_reward") == "True":
                        reward_logs = LogProcessor.fetch_reward_logs()
                        rerun_file = Path(path + "/reward.log")
                        LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                else:
                    if Base_Actions.is_log_capture_required("bool_capt_log_different_files") == "True":
                        testCaseID = str(item.nodeid).split('/')
                        finalTestCaseID = testCaseID[len(testCaseID) - 1]
                        logFileName = str(finalTestCaseID).split('::')[1]

                        path = DirectoryCreator.getDirectoryPath("ServerLog") + "/" + logFileName
                        Path(path).mkdir(parents=True, exist_ok=True)
                        print("Fetching Logs first time")

                        TCIdWithTimeStamp = str(item.nodeid) + '_' + str(datetime.now().time())

                        if GlobalVariables.api_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_api") == "True":
                            apiLogs = LogProcessor.fetch_API_logs()
                            rerun_file = Path(path + "/api.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, apiLogs)

                        if GlobalVariables.portal_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_portal") == "True":
                            portalLogs = LogProcessor.fetch_portal_logs()
                            rerun_file = Path(path + "/portal.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, portalLogs)

                        if GlobalVariables.middleware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_middleware") == "True":
                            mwareLogs = LogProcessor.fetch_middleware_logs()
                            rerun_file = Path(path + "/middleware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, mwareLogs)

                        if GlobalVariables.cnpware_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_cnpware") == "True":
                            cnpWareLogs = LogProcessor.fetch_cnpware_logs()
                            rerun_file = Path(path + "/cnpware.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, cnpWareLogs)

                        if GlobalVariables.config_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_config") == "True":
                            config_logs = LogProcessor.fetch_config_logs()
                            rerun_file = Path(path + "/config.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, config_logs)

                        if GlobalVariables.closedloop_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_closedloop") == "True":
                            closedloop_logs = LogProcessor.fetch_closed_loop_logs()
                            rerun_file = Path(path + "/closedloop.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, closedloop_logs)

                        if GlobalVariables.q2_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_q2") == "True":
                            q2_logs = LogProcessor.fetch_q2_logs()
                            rerun_file = Path(path + "/q2.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, q2_logs)

                        if GlobalVariables.commx_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_commx") == "True":
                            commxLogs = LogProcessor.fetch_commx_logs()
                            rerun_file = Path(path + "/commx.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, commxLogs)

                        if GlobalVariables.ezestore_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_ezestore") == "True":
                            ezestoreLogs = LogProcessor.fetch_ezestore_logs()
                            rerun_file = Path(path + "/ezestore.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, ezestoreLogs)

                        if GlobalVariables.khata_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_khata") == "True":
                            khata_logs = LogProcessor.fetch_khata_logs()
                            rerun_file = Path(path + "/khata.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, khata_logs)

                        if GlobalVariables.reward_logs and Base_Actions.is_log_capture_required(
                                "bool_capt_log_reward") == "True":
                            reward_logs = LogProcessor.fetch_reward_logs()
                            rerun_file = Path(path + "/reward.log")
                            LogProcessor.appendLogs(rerun_file, TCIdWithTimeStamp, reward_logs)
                    else:
                        print("Both bool_capt_log_one_file and bool_capt_log_different_files are disabled")

    GlobalVariables.time_calc.log_collection.pause()
    print(colored("Log Collection Timer paused in 'log on sucess' function in conftest".center(shutil.get_terminal_size().columns, "="), 'cyan'))

    GlobalVariables.time_calc.log_collection.end()
    print(colored("Log Collection Timer ended in 'fin' of method_setup fixture".center(shutil.get_terminal_size().columns, "="),'cyan'))

    GlobalVariables.time_calc.save()
    GlobalVariables.time_calc = None
    print(colored("Saved time_calc object in 'fin' of method_setup fixture".center(shutil.get_terminal_size().columns, "="),'cyan'))


def pytest_sessionstart(session):
    print("Session setup level")
    TestSuiteSetup.ssh_connection(router_ip, router_port, router_username, key_filename)


def pytest_sessionfinish(session, exitstatus):
    print("Session teardown level")
    updateExcel_With_Deselect_And_Broken()

    # Added on Apr 11
    updateExcel_With_Category_And_Subcategory()
    updateExcel_With_RerunAttempts()

    ssh.close()