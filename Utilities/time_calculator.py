import time
import os
import sys
import pickle
import shutil
from datetime import datetime

import pandas as pd
import warnings
import openpyxl
from termcolor import colored

from DataProvider.GlobalConstants import EZEAUTO_MAIN_DIR
from Utilities import DirectoryCreator

from Utilities.ExcelProcessor import getColumnNumberFromName as get_column_number_from_name
from Utilities.ExcelProcessor import getRowNumberFromValue as get_row_number_from_value

warnings.filterwarnings('ignore')

DYNAMIC_EXCEL_REPORT_PATH = DirectoryCreator.getDirectoryPath("ExcelReport") + "/Report.xlsx"

RUNTIME_DIR = os.path.join(EZEAUTO_MAIN_DIR, 'Runtime')
RUNTIME_EXCEL_FILE_URL = os.path.join(RUNTIME_DIR, "TIME_CALCULATOR.xlsx")
TIME_CALC_OBJECTS_DIR = ".time_calculator_objects"
FULL_TIME_CALC_OBJECTS_DIR = os.path.join(RUNTIME_DIR, TIME_CALC_OBJECTS_DIR)

OBJECT_FILENAME_SPACE = '___'  # 3 underscores for now



class BaseTimeCalculator:

    def __init__(self, ) -> None:
        self._start_time_sec = None
        self._pause_time_sec = None
        # self._end_time_sec = None
        self._time_taken = None
        self.is_started = False
        self.is_paused = False
        self.counter_time = 0
        self.counter_time_list = []

    def start(self):
        self._start_time_sec = time.monotonic()
        self.is_started = True

    def pause(self):
        if self.is_started:
            self._pause_time_sec = time.monotonic()
            _time_taken_btw_current_start_n_pause = self._pause_time_sec - self._start_time_sec

            # adding time to the counter list for later summing up
            self.counter_time_list.append(_time_taken_btw_current_start_n_pause)
            self.is_paused = True
            self._start_time_sec = None
        else:
            raise Exception("The time counter has not yet started")

    def resume(self):
        if self.is_started and self.is_paused:
            self._start_time_sec = time.monotonic()
            self.is_paused = False
        elif self.is_started and (not self.is_paused):
            raise Exception("The timer has started but not yet paused")
        else:
            raise Exception("It looks like counter has not yet started")

    def end(self):
        if not self.is_paused:
            self.pause()
        self._time_taken = sum(self.counter_time_list)

    def get_time_taken(self):
        return self._time_taken

    def __str__(self):
        return f"Time taken by the counter is {self._time_taken}" if self._time_taken else "Time counter is not performed"

    def __repr__(self):
        return f"TimeCalculator: {self.__str__()}"

    def __call__(self):
        return self._time_taken


class ExecutionTimeCalculator(BaseTimeCalculator):
    _name = "Execution"

    def __init__(self):
        super().__init__()

    def get_name(self):
        return self._name


class ValidationTimeCalculator(BaseTimeCalculator):
    _name = "Validation"

    def __init__(self):
        super().__init__()

    def get_name(self):
        return self._name


class LogCollectionTimeCalculator(BaseTimeCalculator):
    _name = "Log Collection"

    def __init__(self):
        super().__init__()

    def get_name(self):
        return self._name


class SetupTimeCalculator(BaseTimeCalculator):
    def __init__(self):
        super().__init__()


class TeardownTimeCalculator(BaseTimeCalculator):
    def __init__(self):
        super().__init__()


class EzeAutoTimeCalculator:
    """This class is used to capture and calculate time of test-cases.\n
    Usage: 
        calc = EzeAutoTimeCalculator()\n
        \n
        calc.execution.start()\n
        # do the execution\n
        calc.execution.end()\n
        \n
        calc.validation.start()\n
        # do the validation\n
        calc.validation.end()\n
        \n
        calc.log_collection.start()\n
        # do the log_collection\n
        calc.log_collection.end()\n
        \n
        calc.save()
    """

    _time_format = "%Y-%m-%d %H:%M:%S"

    def __init__(self):
        self._get_testcase_id_n_testfile_relative_filepath()
        self._get_save_object_path()
        # self.load_previous_retries()

        self.setup = SetupTimeCalculator()
        self.execution = ExecutionTimeCalculator()
        self.validation = ValidationTimeCalculator()
        self.log_collection = LogCollectionTimeCalculator()
        self.teardown = TeardownTimeCalculator()
        self.previous_attempts = []

    def _get_testcase_id_n_testfile_relative_filepath(self):
        self.current_testcase_loc_n_test_function_name = os.getenv('PYTEST_CURRENT_TEST')
        print("TC_INFO:", self.current_testcase_loc_n_test_function_name)

        self.testfile_relative_filepath, function_name_info = self.current_testcase_loc_n_test_function_name.split("::")
        self.testcase_id = function_name_info.split(" (")[0]  # not ideal

        print(colored(f" THE TEST CASE INFO: {self.testfile_relative_filepath}:'{self.testcase_id}' "
              .center(shutil.get_terminal_size().columns, "|"), 'yellow'))

    def _get_save_object_path(self):
        if not os.path.isdir(FULL_TIME_CALC_OBJECTS_DIR):
            os.mkdir(FULL_TIME_CALC_OBJECTS_DIR)
        self.save_object_path = os.path.join(
            FULL_TIME_CALC_OBJECTS_DIR,
            f"{self.testcase_id}{OBJECT_FILENAME_SPACE}\
{self.testfile_relative_filepath.lstrip('/').replace('/', '--->')}\
{OBJECT_FILENAME_SPACE}ezeauto_time_calculator.list")  # this is for retries

    def _get_total_time_for_current_attempt(self):
        ls1 = [self.setup(), self.execution(), self.validation(), self.log_collection(), self.teardown()]
        time_list = [i for i in ls1 if i]
        if time_list:
            return sum(time_list)
        else:
            raise Exception("Cannot calculate sum. Please check if all counters are started and ended")

    def load_previous_retries(self):
        if os.path.isfile(self.save_object_path):
            with open(self.save_object_path, "rb") as f:
                self.previous_attempts = pickle.load(f)

    def save(self):
        save_dict = dict(
            testcase_id=self.testcase_id,
            save_time=datetime.strftime(datetime.now(), self._time_format),
            setup=self.setup(),
            execution=self.execution(),
            validation=self.validation(),
            log_collection=self.log_collection(),
            teardown=self.teardown(),
            total_time=self._get_total_time_for_current_attempt(),
            testfile_path=self.testfile_relative_filepath,
            current_testcase_info=self.current_testcase_loc_n_test_function_name
        )
        self.load_previous_retries()
        self.previous_attempts.append(save_dict)
        with open(self.save_object_path, 'wb') as f:
            pickle.dump(self.previous_attempts, f)

        print(colored(shutil.get_terminal_size().columns *"|", 'yellow'))


def generate_timing_dataframe_for_all_attempts_of_testcases():
    if os.path.isdir(FULL_TIME_CALC_OBJECTS_DIR):
        testcase_time_files = [f for f in os.listdir(FULL_TIME_CALC_OBJECTS_DIR) if f.endswith('.list')]
    else:
        testcase_time_files = []

    dfs = []
    if testcase_time_files:
        for i, testcase_time_file in enumerate(testcase_time_files, start=1):
            tc_time_file_full_path = os.path.join(FULL_TIME_CALC_OBJECTS_DIR, testcase_time_file)
            with open(tc_time_file_full_path, 'rb') as f:
                records = pickle.load(f)
            df = pd.DataFrame(records)
            df['unique_tc_id'] = i
            df['save_time'] = pd.to_datetime(df['save_time'])
            df['attempts'] = df['save_time'].rank().astype(int)
            df.sort_values(by='attempts', inplace=True)
            dfs.append(df)

    columns = [
        'unique_tc_id', 'testfile_path', 'testcase_id', 'attempts', 'save_time',
        'setup', 'execution', 'validation', 'log_collection', 'teardown',
        'total_time', 'current_testcase_info',
    ]
    if dfs:
        df = pd.concat(dfs).reset_index(drop=True)[columns]
    else:
        df = pd.DataFrame(columns=columns)
    return df


def process_all_testcase_attempts_dataframe_to_get_needed_data_only(df):
    columns_to_drop = ['execution', 'setup', 'teardown', 'total_time']

    xl = df.copy()
    xl['total_execution_time'] = xl[['setup', 'execution', 'teardown']].sum(axis=1)

    order_of_columns = [
        'unique_tc_id', 'testfile_path', 'testcase_id', 'attempts', 'save_time',
        'total_execution_time', 'validation', 'log_collection',
        'current_testcase_info',
    ]

    xl.drop(columns=columns_to_drop, inplace=True)

    xl = xl[order_of_columns]
    print(colored(" ORIGINAL TIME CALCULATIONS ".center(shutil.get_terminal_size().columns, "="), 'green'))
    print(xl[["testcase_id",'attempts','save_time','total_execution_time','validation','log_collection', 'current_testcase_info']])
    print(colored(shutil.get_terminal_size().columns *"=", 'green'))

    # reducing to one row for each testcase 
    xl = xl.groupby('unique_tc_id').aggregate({
        'testfile_path': pd.Series.mode,
        'testcase_id': pd.Series.mode,
        'current_testcase_info': pd.Series.mode,
        'attempts': 'max',
        'total_execution_time': 'sum',
        'validation': 'sum',
        'log_collection': 'sum',
    })

    xl.reset_index(inplace=True)  # unpivot

    # calculating total time for each testcase
    xl['total_time'] = xl[['total_execution_time', 'validation', 'log_collection']].sum(axis=1)

    return xl


def generate_excel_report(remove_temp_dir_after_generating_report=True):
    df = generate_timing_dataframe_for_all_attempts_of_testcases()
    xl = process_all_testcase_attempts_dataframe_to_get_needed_data_only(df)

    # column mapping
    xl_col_map = dict(
        total_execution_time="Execution Time (sec)",
        validation="Validation Time (sec)",
        log_collection="Log Coll Time (sec)",
        total_time="Total Time (sec)",
    )

    xl.rename(columns=xl_col_map, inplace=True)

    # writing to excel file to runtime dir
    writer = pd.ExcelWriter(RUNTIME_EXCEL_FILE_URL, engine='xlsxwriter')
    xl.to_excel(writer, sheet_name='reduced_report', index=False)
    df.to_excel(writer, sheet_name="detailed_report", index=False)
    writer.save()

    # clearing the temporary files from RUNTIME DIR
    if remove_temp_dir_after_generating_report:
        shutil.rmtree(FULL_TIME_CALC_OBJECTS_DIR)
        os.mkdir(FULL_TIME_CALC_OBJECTS_DIR)

    # ================== MAPPING VALUES TO FRAMEWORK's REPORT FILE =========================
    xl.set_index("testcase_id", inplace=True)  # setting index for .loc purpose of pandas

    workbook = openpyxl.load_workbook(DYNAMIC_EXCEL_REPORT_PATH)
    # print(pd.read_excel(DYNAMIC_EXCEL_REPORT_PATH))
    sheet = workbook["Sheet1"]
    for testcase_id in xl.index.unique():
        print("Updating Excel sheet timings for TC_ID:", testcase_id)
        row_number = get_row_number_from_value(workbook, sheet, 'Test Case ID', testcase_id)
        for column_name in xl_col_map.values():  # float("{:.2f}".format(x))
            value_to_be_set = xl.loc[testcase_id, column_name]  # .iloc[0]  # getting from my df
            column_number = get_column_number_from_name(workbook, sheet, column_name)
            sheet.cell(row=row_number, column=column_number).value = float("{:.2f}".format(value_to_be_set))

    workbook.save(DYNAMIC_EXCEL_REPORT_PATH)
    workbook.close()

    print(colored(" GENERAL TIME CALCULATIONS ".center(shutil.get_terminal_size().columns, "="), 'green'))
    print(xl[["Execution Time (sec)", "Validation Time (sec)", "Log Coll Time (sec)", "Total Time (sec)"]])
    print(colored(shutil.get_terminal_size().columns *"=", 'green'))

    # ========================================================================================

    return xl


# class BaseTimeCalculator:

#     def __init__(self, ) -> None:
#         self._start_time_sec = None
#         self._pause_time_sec = None
#         # self._end_time_sec = None
#         self._time_taken = None
#         self.is_started = False
#         self.is_paused = False
#         self.counter_time = 0
#         self.counter_time_list = []

#     def start(self):
#         self._start_time_sec = time.monotonic()
#         self.is_started = True

#     def pause(self):
#         if self.is_started:
#             self._pause_time_sec = time.monotonic()
#             _time_taken_btw_current_start_n_pause = self._pause_time_sec - self._start_time_sec

#             # adding time to the counter list for later summing up
#             self.counter_time_list.append(_time_taken_btw_current_start_n_pause)
#             self.is_paused = True
#             self._start_time_sec = None
#         else:
#             raise Exception("The time counter has not yet started")

#     def resume(self):
#         if self.is_started and self.is_paused:
#             self._start_time_sec = time.monotonic()
#             self.is_paused = False
#         elif self.is_started and (not self.is_paused):
#             raise Exception("The timer has started but not yet paused")
#         else:
#             raise Exception("It looks like counter has not yet started")

#     def end(self):
#         if not self.is_paused:
#             self.pause()
#         self._time_taken = sum(self.counter_time_list)

#     def get_time_taken(self):
#         return self._time_taken

#     def __str__(self):
#         return f"Time taken by the counter is {self._time_taken}" if self._time_taken else "Time counter is not performed"

#     def __repr__(self):
#         return f"TimeCalculator: {self.__str__()}"

#     def __call__(self):
#         return self._time_taken


# class ExecutionTimeCalculator(BaseTimeCalculator):
#     _name = "Execution"

#     def __init__(self):
#         super().__init__()

#     def get_name(self):
#         return self._name


# class ValidationTimeCalculator(BaseTimeCalculator):
#     _name = "Validation"

#     def __init__(self):
#         super().__init__()

#     def get_name(self):
#         return self._name


# class LogCollectionTimeCalculator(BaseTimeCalculator):
#     _name = "Log Collection"

#     def __init__(self):
#         super().__init__()

#     def get_name(self):
#         return self._name


# class SetupTimeCalculator(BaseTimeCalculator):
#     def __init__(self):
#         super().__init__()


# class TeardownTimeCalculator(BaseTimeCalculator):
#     def __init__(self):
#         super().__init__()


# class EzeAutoTimeCalculator:
#     """This class is used to capture and calculate time of test-cases.\n
#     Usage: 
#         calc = EzeAutoTimeCalculator()\n
#         \n
#         calc.execution.start()\n
#         # do the execution\n
#         calc.execution.end()\n
#         \n
#         calc.validation.start()\n
#         # do the validation\n
#         calc.validation.end()\n
#         \n
#         calc.log_collection.start()\n
#         # do the log_collection\n
#         calc.log_collection.end()\n
#         \n
#         calc.save()
#     """

#     _time_format = "%Y-%m-%d %H:%M:%S"

#     def __init__(self):
#         self._get_testcase_id_n_testfile_relative_filepath()
#         self._get_save_object_path()
#         # self.load_previous_retries()

#         self.setup = SetupTimeCalculator()
#         self.execution = ExecutionTimeCalculator()
#         self.validation = ValidationTimeCalculator()
#         self.log_collection = LogCollectionTimeCalculator()
#         self.teardown = TeardownTimeCalculator()
#         self.previous_attempts = []

#     def _get_testcase_id_n_testfile_relative_filepath(self):
#         self.current_testcase_loc_n_test_function_name = os.getenv('PYTEST_CURRENT_TEST')
#         print("------> ", self.current_testcase_loc_n_test_function_name)

#         self.testfile_relative_filepath, function_name_info = self.current_testcase_loc_n_test_function_name.split("::")
#         self.testcase_id = function_name_info.split(" (")[0]  # not ideal

#         print(f" THE TEST CASE INFO: {self.testfile_relative_filepath}:'{self.testcase_id}' "
#               .center(shutil.get_terminal_size().columns, "|"))

#     def _get_save_object_path(self):
#         if not os.path.isdir(FULL_TIME_CALC_OBJECTS_DIR):
#             os.mkdir(FULL_TIME_CALC_OBJECTS_DIR)
#         self.save_object_path = os.path.join(
#             FULL_TIME_CALC_OBJECTS_DIR,
#             f"{self.testcase_id}{OBJECT_FILENAME_SPACE}\
# {self.testfile_relative_filepath.lstrip('/').replace('/', '--->')}\
# {OBJECT_FILENAME_SPACE}ezeauto_time_calculator.list")  # this is for retries

#     def _get_total_time_for_current_attempt(self):
#         ls1 = [self.setup(), self.execution(), self.validation(), self.log_collection(), self.teardown()]
#         time_list = [i for i in ls1 if i]
#         if time_list:
#             return sum(time_list)
#         else:
#             raise Exception("Cannot calculate sum. Please check if all counters are started and ended")

#     def load_previous_retries(self):
#         if os.path.isfile(self.save_object_path):
#             with open(self.save_object_path, "rb") as f:
#                 self.previous_attempts = pickle.load(f)

#     def save(self):
#         save_dict = dict(
#             testcase_id=self.testcase_id,
#             save_time=datetime.strftime(datetime.now(), self._time_format),
#             setup=self.setup(),
#             execution=self.execution(),
#             validation=self.validation(),
#             log_collection=self.log_collection(),
#             teardown=self.teardown(),
#             total_time=self._get_total_time_for_current_attempt(),
#             testfile_path=self.testfile_relative_filepath,
#             current_testcase_info=self.current_testcase_loc_n_test_function_name
#         )
#         self.load_previous_retries()
#         self.previous_attempts.append(save_dict)
#         with open(self.save_object_path, 'wb') as f:
#             pickle.dump(self.previous_attempts, f)


# def generate_timing_dataframe_for_all_attempts_of_testcases():
#     if os.path.isdir(FULL_TIME_CALC_OBJECTS_DIR):
#         testcase_time_files = [f for f in os.listdir(FULL_TIME_CALC_OBJECTS_DIR) if f.endswith('.list')]
#     else:
#         testcase_time_files = []

#     dfs = []
#     if testcase_time_files:
#         for i, testcase_time_file in enumerate(testcase_time_files, start=1):
#             tc_time_file_full_path = os.path.join(FULL_TIME_CALC_OBJECTS_DIR, testcase_time_file)
#             with open(tc_time_file_full_path, 'rb') as f:
#                 records = pickle.load(f)
#             df = pd.DataFrame(records)
#             df['unique_tc_id'] = i
#             df['save_time'] = pd.to_datetime(df['save_time'])
#             df['attempts'] = df['save_time'].rank().astype(int)
#             df.sort_values(by='attempts', inplace=True)
#             dfs.append(df)

#     columns = [
#         'unique_tc_id', 'testfile_path', 'testcase_id', 'attempts', 'save_time',
#         'setup', 'execution', 'validation', 'log_collection', 'teardown',
#         'total_time', 'current_testcase_info',
#     ]
#     if dfs:
#         df = pd.concat(dfs).reset_index(drop=True)[columns]
#     else:
#         df = pd.DataFrame(columns=columns)
#     return df


# def process_all_testcase_attempts_dataframe_to_get_needed_data_only(df):
#     columns_to_drop = ['execution', 'setup', 'teardown', 'total_time']

#     xl = df.copy()
#     xl['total_execution_time'] = xl[['setup', 'execution', 'teardown']].sum(axis=1)

#     order_of_columns = [
#         'unique_tc_id', 'testfile_path', 'testcase_id', 'attempts', 'save_time',
#         'total_execution_time', 'validation', 'log_collection',
#         'current_testcase_info',
#     ]

#     xl.drop(columns=columns_to_drop, inplace=True)

#     xl = xl[order_of_columns]

#     # reducing to one row for each testcase 
#     xl = xl.groupby('unique_tc_id').aggregate({
#         'testfile_path': pd.Series.mode,
#         'testcase_id': pd.Series.mode,
#         'current_testcase_info': pd.Series.mode,
#         'attempts': 'max',
#         'total_execution_time': 'sum',
#         'validation': 'sum',
#         'log_collection': 'sum',
#     })

#     xl.reset_index(inplace=True)  # unpivot

#     # calculating total time for each testcase
#     xl['total_time'] = xl[['total_execution_time', 'validation', 'log_collection']].sum(axis=1)

#     return xl


# def generate_excel_report(remove_temp_dir_after_generating_report=False):
#     df = generate_timing_dataframe_for_all_attempts_of_testcases()
#     xl = process_all_testcase_attempts_dataframe_to_get_needed_data_only(df)

#     # column mapping
#     xl_col_map = dict(
#         total_execution_time="Execution Time (sec)",
#         validation="Validation Time (sec)",
#         log_collection="Log Coll Time (sec)",
#         total_time="Total Time (sec)",
#     )

#     xl.rename(columns=xl_col_map, inplace=True)

#     # writing to excel file to runtime dir
#     writer = pd.ExcelWriter(RUNTIME_EXCEL_FILE_URL, engine='xlsxwriter')
#     xl.to_excel(writer, sheet_name='reduced_report', index=False)
#     df.to_excel(writer, sheet_name="detailed_report", index=False)
#     writer.save()

#     # clearing the temporary files from RUNTIME DIR
#     if remove_temp_dir_after_generating_report:
#         shutil.rmtree(FULL_TIME_CALC_OBJECTS_DIR)
#         os.mkdir(FULL_TIME_CALC_OBJECTS_DIR)

#     # ================== MAPPING VALUES TO FRAMEWORK's REPORT FILE =========================
#     xl.set_index("testcase_id", inplace=True)  # setting index for .loc purpose of pandas

#     workbook = openpyxl.load_workbook(DYNAMIC_EXCEL_REPORT_PATH)
#     sheet = workbook["Sheet1"]
#     for testcase_id in xl.index.unique():
#         row_number = get_row_number_from_value(workbook, sheet, 'Test Case ID', testcase_id)
#         for column_name in xl_col_map.values():  # float("{:.2f}".format(x))
#             value_to_be_set = xl.loc[testcase_id, column_name]  # .iloc[0]  # getting from my df
#             column_number = get_column_number_from_name(workbook, sheet, column_name)
#             sheet.cell(row=row_number, column=column_number).value = float("{:.2f}".format(value_to_be_set))

#     workbook.save(DYNAMIC_EXCEL_REPORT_PATH)
#     workbook.close()

#     # ========================================================================================

#     return xl
