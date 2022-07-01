from datetime import datetime

import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from openpyxl.styles import PatternFill, Font, Side, Border
from prettytable import PrettyTable
import DataProvider.GlobalConstants
from DataProvider import GlobalVariables
from PageFactory import Base_Actions
from Utilities import ExcelProcessor
from Utilities import ConfigReader, Rerun
from Utilities import DirectoryCreator
from Utilities.execution_log_processor import EzeAutoLogger
logger = EzeAutoLogger(__name__)


EXCEL_reportFilePath = DirectoryCreator.getDirectoryPath("ExcelReport")+"/Report.xlsx"

def get_TC_Exe_Time():
    current = datetime.now()
    GlobalVariables.EXCEL_TC_Exe_completed_time = current.strftime("%H:%M:%S")
    FMT = '%H:%M:%S'
    try:
        totalExecutionTime = datetime.strptime(GlobalVariables.EXCEL_TC_Exe_completed_time, FMT) - datetime.strptime(str(
        GlobalVariables.EXCEL_TC_Exe_Starting_Time), FMT)
    except Exception as e:
        print("Unable to set the totalExecutionTime due to error : " + str(e) + ". Hence setting it to 0.")
        totalExecutionTime = 0

    # Converting time duration to seconds
    GlobalVariables.EXCEL_Execution_Time = sum(x * int(t) for x, t in zip([3600, 60, 1], str(totalExecutionTime).split(":")))

    # Converting time duration to milliseconds
    # GlobalVariables.EXCEL_Execution_Time = Exe_Time_Sec * 1000


def get_TC_Val_Time():
    current = datetime.now()
    GlobalVariables.EXCEL_TC_Val_completed_time = current.strftime("%H:%M:%S")
    print("TC Val completed time:", GlobalVariables.EXCEL_TC_Val_completed_time)
    FMT = '%H:%M:%S'
    try:
        totalValidationTime = datetime.strptime(GlobalVariables.EXCEL_TC_Val_completed_time, FMT) - datetime.strptime(str(
        GlobalVariables.EXCEL_TC_Val_Starting_Time), FMT)
    except Exception as e:
        print("Unable to set the totalValidationTime due to error : " + str(e) + ". Hence setting it to 0.")
        totalValidationTime = 0

    # Converting time duration to seconds
    GlobalVariables.EXCEL_Val_time = sum(x * int(t) for x, t in zip([3600, 60, 1], str(totalValidationTime).split(":")))

    # Converting time duration to milliseconds
    # GlobalVariables.EXCEL_Val_time = Val_Time_Sec * 1000


def get_Log_Collection_Time():
    current = datetime.now()
    GlobalVariables.EXCEL_TC_Val_completed_time = current.strftime("%H:%M:%S")
   # GlobalVariables.EXCEL_TC_Val_completed_time = current.strftime("%H:%M:%S")
    FMT = '%H:%M:%S'
    print("Log collection end time: ", GlobalVariables.EXCEL_TC_Val_completed_time)
    try:
        totalValidationTime = datetime.strptime(GlobalVariables.EXCEL_TC_Val_completed_time, FMT) - datetime.strptime(str(
        GlobalVariables.EXCEL_TC_LogColl_Starting_Time), FMT)
    except Exception as e:
        print("Unable to set the totalValidationTime due to error : " + str(e) + ". Hence setting it to 0.")
        totalValidationTime = 0

    # Converting time duration to seconds
    GlobalVariables.EXCEL_LogCollTime = sum(x * int(t) for x, t in zip([3600, 60, 1], str(totalValidationTime).split(":")))

    # Converting time duration to milliseconds
    # GlobalVariables.EXCEL_LogCollTime = Val_Time_Sec * 1000

def createStatusTable():
    apiVal = GlobalVariables.str_api_val_result
    dbVal = GlobalVariables.str_db_val_result
    portalVal = GlobalVariables.str_portal_val_result
    appVal = GlobalVariables.str_app_val_result
    uiVal = GlobalVariables.str_ui_val_result

    get_TC_Val_Time()
    print("portalVal: ", portalVal)
    print("")
    print("")

    # VALIDATION TABLE DETAILS
    myTable = PrettyTable(["Validation Type", "Validation Status"])
    myTable.title = 'Validation Details'

    if apiVal == "Failed":
        myTable.add_row(["API Validation", "Fail"])
    else:
        myTable.add_row(["API Validation", apiVal])

    if dbVal == "Failed":
        myTable.add_row(["DB Validation", "Fail"])
    else:
        myTable.add_row(["DB Validation", dbVal])

    if portalVal == "Failed":
        myTable.add_row(["Portal Validation", "Fail"])
    else:
        myTable.add_row(["Portal Validation", portalVal])

    if appVal == "Failed":
        myTable.add_row(["App Validation", "Fail"])
    else:
        myTable.add_row(["App Validation", appVal])

    if uiVal == "Failed":
        myTable.add_row(["UI Validation", "Fail"])
    else:
        myTable.add_row(["UI Validation", uiVal])

    myTable.align = 'l'
    print(myTable)
    print("")
    print("")

    myTable1 = PrettyTable()
    myTable1.title = 'Debugging Info'
    myTable1.header = True
    myTable1.field_names = ["Type", "API", "Middleware", "Cnpware", "Portal", "App"]

    if Base_Actions.is_log_capture_required("bool_capt_log_fail") == "True" or Base_Actions.is_log_capture_required(
            "bool_capt_log_pass") == "True":
        if Base_Actions.is_log_capture_required("bool_capt_log_api") == "True" and GlobalVariables.apiLogs:
            apiLogs = 'Yes'
        else:
            apiLogs = 'No'
        if Base_Actions.is_log_capture_required(
                "bool_capt_log_middleware") == "True" and GlobalVariables.middleWareLogs:
            mWareLogs = 'Yes'
        else:
            mWareLogs = 'No'
        if Base_Actions.is_log_capture_required("bool_capt_log_cnpware") == "True" and GlobalVariables.cnpWareLogs:
            cnpWareLogs = 'Yes'
        else:
            cnpWareLogs = 'No'
        if Base_Actions.is_log_capture_required("bool_capt_log_portal") == "True" and GlobalVariables.portalLogs:
            portalLogs = 'Yes'
        else:
            portalLogs = 'No'
    else:
        apiLogs = 'No'
        mWareLogs = 'No'
        cnpWareLogs = 'No'
        portalLogs = 'No'

    if GlobalVariables.EXCEL_TC_Execution == "Fail" or GlobalVariables.str_api_val_result == "Fail" or GlobalVariables.str_db_val_result == "Fail" or GlobalVariables.str_portal_val_result == "Fail" or GlobalVariables.str_app_val_result == "Fail" or GlobalVariables.str_ui_val_result == "Fail":
        myTable1.add_row(["Log Captured", apiLogs, mWareLogs, cnpWareLogs, portalLogs, "N/A"])
    else:
        myTable1.add_row(["Log Captured", apiLogs, mWareLogs, cnpWareLogs, portalLogs, "N/A"])

    # SCREENSHOT INFO
    appSS = 'No'
    portalSS = 'No'

    if Base_Actions.is_ss_capture_required("bool_capt_ss_pass") == "True":
        if GlobalVariables.bool_ss_portal_val == "Passed" or portalVal == 'Pass':
            portalSS = 'Yes'
        if GlobalVariables.bool_ss_app_val == "Passed" or appVal == 'Pass':
            appSS = 'Yes'

    if Base_Actions.is_ss_capture_required("bool_capt_ss_fail") == "True":
        if GlobalVariables.bool_ss_portal_val == "Failed" or portalVal == 'Failed':
            portalSS = 'Yes'

        if GlobalVariables.bool_ss_app_val == "Failed" or appVal == 'Failed':
            appSS = 'Yes'

    myTable1.add_row(["Screenshot Captured", "N/A", "N/A", "N/A", portalSS, appSS])
    myTable1.align = 'l'

    print(myTable1)
    print("")



def revert_excel_global_variables():
    GlobalVariables.EXCEL_TC_Exe_Starting_Time = 00
    GlobalVariables.EXCEL_TC_Exe_completed_time = 00

    GlobalVariables.EXCEL_TC_Val_Starting_Time = 00
    GlobalVariables.EXCEL_TC_Val_completed_time = 00

    GlobalVariables.EXCEL_TC_LogColl_Starting_Time = 00
    GlobalVariables.EXCEL_TC_LogColl_completed_time = 00
    GlobalVariables.EXCEL_TC_Execution = "Skip"
    GlobalVariables.str_api_val_result = "N/A"
    GlobalVariables.str_db_val_result = "N/A"
    GlobalVariables.str_portal_val_result = "N/A"
    GlobalVariables.str_app_val_result = "N/A"
    GlobalVariables.str_ui_val_result = "N/A"
    GlobalVariables.bool_chargeslip_val_result = "N/A"
    # GlobalVariables.apiLogs = False
    # GlobalVariables.portalLogs = False
    # GlobalVariables.cnpWareLogs = False
    # GlobalVariables.middleWareLogs = False


def setStylesForExcel():
    wb = openpyxl.load_workbook(EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    max_row = sheet.max_row

    colNum_overall = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
    colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
    colNum_apiVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
    colNum_dbVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
    colNum_portalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
    colNum_appVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
    colNum_uiVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')

    column_list = [colNum_overall, colNum_execution, colNum_apiVal, colNum_dbVal, colNum_portalVal, colNum_appVal, colNum_uiVal]

    for column in column_list:
        for row in range(2, max_row + 1):
            if sheet.cell(row, column).value == "Pass":
                sheet.cell(row, column).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif sheet.cell(row, column).value == "Fail":
                sheet.cell(row, column).fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

    for i in range(2, sheet.max_row + 1):
        colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
        if sheet.cell(row=i, column=colNum_overallStatus).value == "Broken":
            sheet.cell(row=i, column=colNum_overallStatus).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        if sheet.cell(row=i, column=colNum_overallStatus).value == "Deselected":
            sheet.cell(row=i, column=colNum_overallStatus).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')



    # Set width for all cells
    sheet.column_dimensions['A'].width = 70 #Test Case ID
    sheet.column_dimensions['B'].width = 15 #File Name
    sheet.column_dimensions['C'].width = 18 #Directory Name
    sheet.column_dimensions['D'].width = 18 #Category
    sheet.column_dimensions['E'].width = 18 #Sub-Category
    sheet.column_dimensions['F'].width = 18 #OverAll Results
    sheet.column_dimensions['G'].width = 18 #TC Execution
    sheet.column_dimensions['H'].width = 15 #API Val
    sheet.column_dimensions['I'].width = 15 #DB Val
    sheet.column_dimensions['J'].width = 15 #Portal Val
    sheet.column_dimensions['K'].width = 15 #App Val
    sheet.column_dimensions['L'].width = 15 #UI Val
    sheet.column_dimensions['M'].width = 22 #Execution Time (sec)
    sheet.column_dimensions['N'].width = 22 #Validation Time (sec)
    sheet.column_dimensions['O'].width = 22 #Log Coll Time (sec)
    sheet.column_dimensions['P'].width = 18 #Total Time (sec)
    sheet.column_dimensions['Q'].width = 18 #Rerun Attempts


    # Set background color and font style
    fill_pattern = PatternFill(patternType='solid', fgColor='87CEEB')
    font = Font(size=11, bold=True, color="121103")

    for x in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=x).font = font
        sheet.cell(row=1, column=x).fill = fill_pattern

    # Set border for all the cells
    side = Side(border_style="thin", color="000000")

    border = Border(left=side, right=side, top=side, bottom=side)
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row, column).border = border

    wb.save(EXCEL_reportFilePath)


def updateExcel_With_Deselect_And_Broken():
    wb = openpyxl.load_workbook(EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
        testcase = (sheet.cell(row=i, column=colNum_testcase)).value
        if testcase in GlobalVariables.list_deselected_testcases:

            colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            sheet.cell(row=i, column=colNum_overallStatus).value = "Deselected"

            colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
            sheet.cell(row=i, column=colNum_execution).value = "N/A"

            colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
            sheet.cell(row=i, column=colNum_APIval).value = "N/A"

            colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
            sheet.cell(row=i, column=colNum_DBval).value = "N/A"

            colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
            sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

            colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
            sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

            colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
            sheet.cell(row=i, column=colNum_UIval).value = "N/A"

        else:
            colNum_overallResult = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            cellValue = (sheet.cell(row=i, column=colNum_overallResult)).value
            if cellValue is None:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = "Broken"

                colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
                sheet.cell(row=i, column=colNum_execution).value = "N/A"

                colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
                sheet.cell(row=i, column=colNum_APIval).value = "N/A"

                colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
                sheet.cell(row=i, column=colNum_DBval).value = "N/A"

                colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
                sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

                colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
                sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

                colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
                sheet.cell(row=i, column=colNum_UIval).value = "N/A"

    wb.save(EXCEL_reportFilePath)


def updateExcel_With_RerunAttempts():
    wb = openpyxl.load_workbook(EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    if ConfigReader.read_config("Validations", "bool_rerun_immediately").lower() == "true" and ConfigReader.read_config(
            "Validations", "bool_rerun_at_the_end").lower() == "false":
        colNum_rerunAttempts = ExcelProcessor.getColumnNumberFromName("", sheet, 'Rerun Attempts')
        for i in range(2, sheet.max_row + 1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            rerunCount = Rerun.getRerunCount(testcase)
            if rerunCount >= 0:
                cellValue_rerunAttempts = int(ConfigReader.read_config("Validations", "int_rerun_count")) - rerunCount
                sheet.cell(row=i, column=colNum_rerunAttempts).value = cellValue_rerunAttempts
            else:
                cellValue_rerunAttempts = ConfigReader.read_config("Validations", "int_rerun_count")
                sheet.cell(row=i, column=colNum_rerunAttempts).value = int(cellValue_rerunAttempts)

    # If both bool_rerun_at_the_end and bool_rerun_immediately are disabled, setting value as N/A for Rerun Attempts
    if ConfigReader.read_config("Validations", "bool_rerun_at_the_end").lower() == "false" and ConfigReader.read_config(
            "Validations", "bool_rerun_immediately").lower() == "false":
        colNum_RerunAttempts = ExcelProcessor.getColumnNumberFromName("", sheet, 'Rerun Attempts')
        for i in range(2, sheet.max_row + 1):
            sheet.cell(row=i, column=colNum_RerunAttempts).value = "N/A"

    wb.save(EXCEL_reportFilePath)


def updateExcel_With_Category_And_Subcategory():
    wb = openpyxl.load_workbook(EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        colNum_directoryName = ExcelProcessor.getColumnNumberFromName("", sheet, 'Directory Name')
        directoryName = (sheet.cell(row=i, column=colNum_directoryName)).value

        category = directoryName.split("/")[0]
        subCategory = directoryName.split("/")[1]

        colNum_category = ExcelProcessor.getColumnNumberFromName("", sheet, 'Category')
        colNum_subCategory = ExcelProcessor.getColumnNumberFromName("", sheet, 'Sub-Category')

        cellValue_category = (sheet.cell(row=i, column=colNum_category)).value
        cellValue_subCategory = (sheet.cell(row=i, column=colNum_subCategory)).value

        if cellValue_category is None:
            sheet.cell(row=i, column=colNum_category).value = category

            if cellValue_subCategory is None:
                sheet.cell(row=i, column=colNum_subCategory).value = subCategory

    wb.save(EXCEL_reportFilePath)


def updateTestCaseResult(msg):
    createStatusTable()

    ls_validation_msg = []
    if not GlobalVariables.bool_val_exe:
        if GlobalVariables.str_api_val_result == "Fail":
            ls_validation_msg.append("API validation Failed!!")
        if GlobalVariables.str_db_val_result == "Fail":
            ls_validation_msg.append("DB validation Failed!!")
        if GlobalVariables.str_portal_val_result == "Fail":
            ls_validation_msg.append("PORTAL validation Failed!!")
        if GlobalVariables.str_app_val_result == "Fail":
            ls_validation_msg.append("APP validation Failed!!")
        if GlobalVariables.str_ui_val_result == "Fail":
            ls_validation_msg.append("UI validation Failed!!")
        if not GlobalVariables.str_chargeslip_val_result:
            ls_validation_msg.append("Charge-Slip validation Failed!!")
        if GlobalVariables.str_api_val_result == "Fail" or GlobalVariables.str_db_val_result == "Fail" or GlobalVariables.str_portal_val_result == "Fail" or GlobalVariables.str_app_val_result == "Fail" or GlobalVariables.str_ui_val_result == "Fail" or GlobalVariables.str_chargeslip_val_result == False:
            message = ""
            for validation_msg in ls_validation_msg:
                message = message + "\n" + validation_msg
            pytest.fail(message)


def capture_ss_when_exe_failed():
    if GlobalVariables.appDriver != '' and Base_Actions.is_ss_capture_required(
            "bool_capt_ss_fail") == "True":
        try:
            allure.attach(GlobalVariables.appDriver.get_screenshot_as_png(), name="app_screen",
                          attachment_type=AttachmentType.PNG)
        except Exception as e:
            logger.exception(f"Unable to take screenshot : {e}")

    if GlobalVariables.portalDriver != '' and Base_Actions.is_ss_capture_required(
            "bool_capt_ss_fail") == "True":
        try:
            allure.attach(GlobalVariables.portalDriver.get_screenshot_as_png(), name="portal_page",
                          attachment_type=AttachmentType.PNG)
        except Exception as e:
            logger.exception(f"Unable to take screenshot : {e}")

    if GlobalVariables.charge_slip_driver != '' and Base_Actions.is_ss_capture_required(
            "bool_capt_ss_fail") == "True":
        try:
            allure.attach(GlobalVariables.charge_slip_driver.get_screenshot_as_png(), name="chargeslip",
                          attachment_type=AttachmentType.PNG)
        except Exception as e:
            logger.exception(f"Unable to take screenshot : {e}")