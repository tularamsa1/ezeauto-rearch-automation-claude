import openpyxl

matchFound = False
global sheet


def initializeWorkbook(workbookPath):
    workbook = openpyxl.load_workbook(workbookPath)
    return workbook


def initializeSheet(workbook, sheetName):
    sheet = workbook[sheetName]
    return sheet


def getColumnNumberFromName(workbook, sheet, columnName):
    matchFound = False
    totalcolumns = sheet.max_column
    i=1
    for col in sheet.iter_cols(0, totalcolumns):
        if col[0].value == columnName:
            matchFound = True
            break
        i+=1
    if matchFound:
        return i
    else:
        print("Entered column value not available. Please verify the Data sheet.")
        return ValueError


def getRowNumberFromValue(workbook, sheet, columnName, rowValue):
    matchFound = False
    totalrows = sheet.max_row
    columnNumber = getColumnNumberFromName(workbook, sheet, columnName)
    i=1
    for currentRow in range(1, totalrows+1):
        if str(sheet.cell(row=currentRow,column=columnNumber).value) == str(rowValue):
            matchFound = True
            break
        i+=1
    if matchFound:
        return i
    else:
        print("Entered row value not available. Please verify the Data sheet.")
        return ValueError


def getCellValue(workbook, sheet, rowValue, ColumnName):
    try:
        columnNumber = getColumnNumberFromName(workbook, sheet, ColumnName)
        rowNumber = getRowNumberFromValue(workbook, sheet, ColumnName, rowValue)
        return sheet.cell(row=rowNumber, column=columnNumber).value
    except ValueError:
        print("Entered row value or column name is not available in the sheet.")


def getAPIEnviURL(workbook, environment):
    sheet = initializeSheet(workbook, "Environment")
    columnNumber = getColumnNumberFromName(workbook, sheet, "URL")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Name", environment)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("Environment details unavailable. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value


def getAPIDetails(workbook, apiName, apidetailRequired):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, apidetailRequired)
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("API|DB details is  not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value


def getAPIValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "API_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("API|DB details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value


def getDBValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "DB_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("DB details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value


def getPortalValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "Portal_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("Portal details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value